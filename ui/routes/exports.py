"""Planning export and month-over-month routes."""

from datetime import datetime
from typing import Callable

from flask import Blueprint, jsonify, request, send_file

from modules.database_exporter import DatabaseExporter
from modules.mom_comparison_engine import MoMComparisonEngine


def _apply_edit_highlights(path: str, engine):
    """Open the exported workbook and apply edit highlights + summary sheet."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.comments import Comment

    # Collect all edits
    all_edits = []
    for _, rows in engine.results.items():
        for row in rows:
            if row.manual_edits:
                for period, edit_data in row.manual_edits.items():
                    original = edit_data.get('original', 0.0)
                    new_val = edit_data.get('new', 0.0)
                    delta_pct = round((new_val - original) / abs(original) * 100, 2) if original != 0 else 0.0
                    all_edits.append({
                        'line_type': row.line_type,
                        'material_number': row.material_number,
                        'material_name': row.material_name,
                        'period': period,
                        'original': original,
                        'new': new_val,
                        'delta_pct': delta_pct,
                    })

    if not all_edits:
        return

    wb = openpyxl.load_workbook(path)
    ws = wb['Planning sheet']

    # Build column lookups from header row
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    period_col = {}
    mat_col_idx = None
    lt_col_idx = None
    for i, val in enumerate(header, start=1):
        if val is None:
            continue
        s = str(val)
        period_col[s] = i
        if s == 'Material number':
            mat_col_idx = i
        elif s == 'Line type':
            lt_col_idx = i

    # Build row lookup: (material_number, line_type) -> row_idx
    row_lookup = {}
    if mat_col_idx and lt_col_idx:
        for row_idx, row_data in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            mat_val = row_data[mat_col_idx - 1]
            lt_val = row_data[lt_col_idx - 1]
            if mat_val and lt_val:
                row_lookup[(str(mat_val), str(lt_val))] = row_idx

    # Fill styles
    yellow_fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
    green_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    red_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
    bold_font = Font(bold=True)

    for edit in all_edits:
        row_idx = row_lookup.get((edit['material_number'], edit['line_type']))
        col_idx = period_col.get(edit['period'])
        if row_idx is None or col_idx is None:
            continue
        cell = ws.cell(row=row_idx, column=col_idx)
        original = edit['original']
        new_val = edit['new']
        delta_pct = edit['delta_pct']
        if new_val > original:
            cell.fill = green_fill
            cell.font = bold_font
        elif new_val < original:
            cell.fill = red_fill
            cell.font = bold_font
        else:
            cell.fill = yellow_fill
        cell.comment = Comment(f"Original: {original}\nNew: {new_val}\nDelta: {delta_pct}%", 'SOP Engine')

    # Edits Summary sheet
    if 'Edits Summary' in wb.sheetnames:
        del wb['Edits Summary']
    ws_edits = wb.create_sheet('Edits Summary')
    ws_edits.append(['Line Type', 'Material Number', 'Material Name', 'Period',
                     'Original Value', 'New Value', 'Delta %'])
    for edit in all_edits:
        ws_edits.append([edit['line_type'], edit['material_number'], edit['material_name'],
                         edit['period'], edit['original'], edit['new'], edit['delta_pct']])

    wb.save(path)


def create_exports_blueprint(
    get_active: Callable[[], tuple],
    export_dir: Callable[[], object],
    cycle_manager: Callable[[], object],
) -> Blueprint:
    bp = Blueprint('exports', __name__)

    @bp.route('/api/export')
    def export():
        _, current_engine = get_active()

        if current_engine is None:
            return jsonify({'error': 'No calculations run'}), 400

        out_dir = export_dir()
        out_dir.mkdir(exist_ok=True)
        export_path = out_dir / f'SOP_Python_Results_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        inventory_quality_engine = None
        try:
            from modules.inventory_quality_engine import InventoryQualityEngine
            inventory_quality_engine = InventoryQualityEngine(
                current_engine.data,
                current_engine.results,
                current_engine.value_results,
            )
        except Exception:
            pass

        previous_df = None
        try:
            cm = cycle_manager()
            if cm.has_previous_cycle():
                previous_df = cm.load_previous_cycle()
                if previous_df.empty:
                    previous_df = None
                    print('[export] Previous cycle loaded but empty - MoM sheet skipped')
                else:
                    print(f'[export] Previous cycle loaded ({len(previous_df)} rows) - MoM sheet will be included')
            else:
                print('[export] No previous cycle on disk - MoM sheet skipped (will be available after next calculation)')
        except Exception as exc:
            print(f'[export] Could not load previous cycle: {exc}')

        current_engine.to_excel_with_values(
            str(export_path),
            inventory_quality_engine=inventory_quality_engine,
            previous_cycle_df=previous_df,
        )
        _apply_edit_highlights(str(export_path), current_engine)

        return send_file(str(export_path), as_attachment=True)

    @bp.route('/api/export_db', methods=['POST'])
    def export_db():
        """Export planning results to a flat DB-ready Excel file via DatabaseExporter."""
        _, current_engine = get_active()
        if current_engine is None:
            return jsonify({'error': 'No calculations run'}), 400

        try:
            planning_df = current_engine.to_dataframe()
            site = getattr(current_engine.data.config, 'site', 'NLX1')
            initial_date = current_engine.data.config.initial_date

            exporter = DatabaseExporter(planning_df, site, initial_date)
            db_df = exporter.export_to_dataframe()

            if db_df.empty:
                return jsonify({'error': 'No data to export (no matching line types)'}), 400

            out_dir = export_dir()
            out_dir.mkdir(exist_ok=True)

            req_data = request.get_json(silent=True) or {}
            filename = req_data.get('filename', '').strip()
            if not filename:
                filename = f'SOP_DB_Export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            safe_name = ''.join(c for c in filename if c.isalnum() or c in '._- ')
            if not safe_name.endswith('.xlsx'):
                safe_name += '.xlsx'

            export_path = out_dir / safe_name
            db_df.to_excel(str(export_path), index=False)
            print(f'[export_db] {len(db_df)} rows written -> {export_path}')

            return send_file(str(export_path), as_attachment=True, download_name=safe_name)
        except Exception as exc:
            import traceback
            return jsonify({'error': str(exc), 'trace': traceback.format_exc()}), 500

    @bp.route('/api/mom')
    def get_mom_comparison():
        """Return sequential period-over-period MoM comparison from the current run."""
        _, current_engine = get_active()
        if current_engine is None:
            return jsonify({'available': False, 'message': 'No calculations run yet. Run calculations first.'})

        try:
            num_months = int(request.args.get('num_months', 6))
            num_months = max(1, min(num_months, 24))

            current_df = current_engine.to_dataframe()
            result = MoMComparisonEngine.calculate_sequential(current_df, num_months)
            return jsonify(result)
        except Exception as exc:
            import traceback
            return jsonify({'error': str(exc), 'trace': traceback.format_exc()}), 500

    return bp
