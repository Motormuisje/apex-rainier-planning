"""Scenario save/load/compare routes."""

import json
import uuid
from collections import defaultdict
from datetime import datetime
from typing import Callable

from flask import Blueprint, jsonify, request, send_file

from modules.models import LineType


def create_scenarios_blueprint(
    scenarios: dict,
    sessions: dict,
    get_active_session_id: Callable[[], str | None],
    get_active: Callable[[], tuple],
    global_config: dict,
    export_dir: Callable[[], object],
    build_pending_edits_from_results_snapshot: Callable[[dict], dict],
    planning_row_from_snapshot: Callable[..., object],
    rebuild_volume_caches_from_results: Callable[[object], None],
    valuation_params_from_config: Callable[[object], object],
    parse_purchased_and_produced: Callable[[object], dict],
    format_purchased_and_produced: Callable[[dict], str],
    row_key_from_obj: Callable[[object], tuple],
) -> Blueprint:
    bp = Blueprint('scenarios', __name__)

    @bp.route('/api/scenarios', methods=['GET'])
    def list_scenarios():
        """List saved scenarios for the active session."""
        active_session_id = get_active_session_id()
        result = [
            {
                'id': sid,
                'name': sc['name'],
                'session_id': sc['session_id'],
                'timestamp': sc['timestamp'],
                'edit_count': sc['edit_count'],
            }
            for sid, sc in scenarios.items()
            if sc['session_id'] == active_session_id
        ]
        result.sort(key=lambda x: x['timestamp'])
        return jsonify({'scenarios': result})

    @bp.route('/api/scenarios/save', methods=['POST'])
    def save_scenario():
        """Deep-copy current volumes + cascaded values into a named scenario."""
        _, current_engine = get_active()
        if current_engine is None:
            return jsonify({'error': 'No calculations run'}), 400

        req = request.get_json() or {}
        name = req.get('name', '').strip()
        if not name:
            return jsonify({'error': 'Scenario name is required'}), 400

        results_snapshot = {}
        total_edits = 0
        for line_type, rows in current_engine.results.items():
            results_snapshot[line_type] = []
            for row in rows:
                results_snapshot[line_type].append({
                    'material_number': row.material_number,
                    'material_name': row.material_name,
                    'product_type': row.product_type,
                    'product_family': row.product_family,
                    'spc_product': row.spc_product,
                    'product_cluster': row.product_cluster,
                    'product_name': row.product_name,
                    'line_type': row.line_type,
                    'aux_column': row.aux_column,
                    'aux_2_column': row.aux_2_column,
                    'starting_stock': row.starting_stock,
                    'values': dict(row.values),
                    'manual_edits': {period: dict(edit) for period, edit in row.manual_edits.items()},
                })
                total_edits += len(row.manual_edits)

        value_snapshot = {}
        for line_type, rows in current_engine.value_results.items():
            value_snapshot[line_type] = []
            for row in rows:
                value_snapshot[line_type].append({
                    'material_number': row.material_number,
                    'material_name': row.material_name,
                    'product_type': row.product_type,
                    'product_family': row.product_family,
                    'spc_product': row.spc_product,
                    'product_cluster': row.product_cluster,
                    'product_name': row.product_name,
                    'line_type': row.line_type,
                    'aux_column': row.aux_column,
                    'aux_2_column': row.aux_2_column,
                    'starting_stock': row.starting_stock,
                    'values': dict(row.values),
                    'manual_edits': {},
                })

        active_session_id = get_active_session_id()
        active_session = sessions.get(active_session_id, {})
        pending_snapshot = json.loads(json.dumps(active_session.get('pending_edits', {})))
        if not pending_snapshot:
            pending_snapshot = build_pending_edits_from_results_snapshot(results_snapshot)

        scenario_id = str(uuid.uuid4())
        scenarios[scenario_id] = {
            'id': scenario_id,
            'name': name,
            'session_id': active_session_id,
            'timestamp': datetime.now().isoformat(),
            'edit_count': total_edits,
            'results': results_snapshot,
            'value_results': value_snapshot,
            'pending_edits': pending_snapshot,
            'value_aux_overrides': json.loads(json.dumps(active_session.get('value_aux_overrides', {}))),
            'valuation_params': {str(k): float(v) for k, v in (global_config.get('valuation_params') or {}).items()},
            'purchased_and_produced': global_config.get('purchased_and_produced', ''),
        }
        return jsonify({'success': True, 'scenario_id': scenario_id, 'name': name, 'edit_count': total_edits})

    @bp.route('/api/scenarios/load', methods=['POST'])
    def load_scenario():
        """Restore a saved scenario into the active engine."""
        sess, current_engine = get_active()
        if current_engine is None:
            return jsonify({'error': 'No calculations run'}), 400

        req = request.get_json() or {}
        scenario_id = req.get('scenario_id', '')
        if not scenario_id or scenario_id not in scenarios:
            return jsonify({'error': 'Scenario not found'}), 404

        active_session_id = get_active_session_id()
        sc = scenarios[scenario_id]
        if sc['session_id'] != active_session_id:
            return jsonify({'error': 'Scenario belongs to a different session'}), 403

        restored_results = {}
        for line_type, snap_rows in (sc.get('results') or {}).items():
            restored_results[line_type] = [
                planning_row_from_snapshot(snap, fallback_line_type=line_type)
                for snap in (snap_rows or [])
            ]
        for line_type in current_engine.results.keys():
            restored_results.setdefault(line_type, [])
        current_engine.results = restored_results

        restored_value_results = {}
        for line_type, snap_rows in (sc.get('value_results') or {}).items():
            restored_value_results[line_type] = [
                planning_row_from_snapshot(snap, fallback_line_type=line_type)
                for snap in (snap_rows or [])
            ]
        for line_type in current_engine.value_results.keys():
            restored_value_results.setdefault(line_type, [])
        current_engine.value_results = restored_value_results

        restored_pending = json.loads(json.dumps(sc.get('pending_edits', {})))
        if not restored_pending:
            restored_pending = build_pending_edits_from_results_snapshot(sc.get('results', {}))
        sess['pending_edits'] = restored_pending
        sess['value_aux_overrides'] = json.loads(json.dumps(sc.get('value_aux_overrides', {})))
        sess['undo_stack'] = []
        sess['redo_stack'] = []
        rebuild_volume_caches_from_results(current_engine)

        sc_vp = sc.get('valuation_params')
        restored_vp = None
        if sc_vp and getattr(current_engine, 'data', None) is not None:
            current_engine.data.valuation_params = valuation_params_from_config(sc_vp)
            global_config['valuation_params'] = {str(k): float(v) for k, v in sc_vp.items()}
            restored_vp = sc_vp

        sc_pap = sc.get('purchased_and_produced')
        if sc_pap is not None and getattr(current_engine, 'data', None) is not None:
            pap_dict = parse_purchased_and_produced(sc_pap) if isinstance(sc_pap, str) else dict(sc_pap)
            current_engine.data.purchased_and_produced = pap_dict
            global_config['purchased_and_produced'] = format_purchased_and_produced(pap_dict)

        results_dict = {
            line_type: [row.to_dict() for row in rows]
            for line_type, rows in current_engine.results.items()
        }
        value_results_dict = {
            line_type: [row.to_dict() for row in rows]
            for line_type, rows in current_engine.value_results.items()
        }
        consolidation = [
            row.to_dict()
            for row in current_engine.value_results.get(LineType.CONSOLIDATION.value, [])
        ]

        baseline = sess.get('reset_baseline')
        baseline_results = None
        baseline_value_results = None
        if baseline:
            if baseline.get('results'):
                baseline_results = {
                    line_type: [{
                        'material_number': row.get('material_number', ''),
                        'aux_column': row.get('aux_column', ''),
                        'values': row.get('values', {}),
                    } for row in rows]
                    for line_type, rows in baseline['results'].items()
                }
            if baseline.get('value_results'):
                baseline_value_results = {
                    line_type: [{
                        'material_number': row.get('material_number', ''),
                        'aux_column': row.get('aux_column', ''),
                        'values': row.get('values', {}),
                    } for row in rows]
                    for line_type, rows in baseline['value_results'].items()
                }

        resp = {
            'success': True,
            'scenario_id': scenario_id,
            'name': sc['name'],
            'results': results_dict,
            'value_results': value_results_dict,
            'consolidation': consolidation,
            'pending_edits': sess.get('pending_edits', {}),
            'value_aux_overrides': sess.get('value_aux_overrides', {}),
        }
        if restored_vp is not None:
            resp['restored_valuation_params'] = restored_vp
        if baseline_results:
            resp['baseline_results'] = baseline_results
        if baseline_value_results:
            resp['baseline_value_results'] = baseline_value_results
        return jsonify(resp)

    @bp.route('/api/scenarios/<scenario_id>', methods=['DELETE'])
    def delete_scenario(scenario_id):
        if scenario_id not in scenarios:
            return jsonify({'error': 'Scenario not found'}), 404
        if scenarios[scenario_id]['session_id'] != get_active_session_id():
            return jsonify({'error': 'Scenario belongs to a different session'}), 403
        del scenarios[scenario_id]
        return jsonify({'success': True})

    @bp.route('/api/scenarios/compare', methods=['POST'])
    def compare_scenarios():
        req = request.get_json() or {}
        id_a = req.get('scenario_a_id', '')
        id_b = req.get('scenario_b_id', '')
        if id_a not in scenarios or id_b not in scenarios:
            return jsonify({'error': 'Scenario not found'}), 404
        sc_a = scenarios[id_a]
        sc_b = scenarios[id_b]
        if sc_a['session_id'] != get_active_session_id() or sc_b['session_id'] != get_active_session_id():
            return jsonify({'error': 'Scenarios belong to a different session'}), 403

        diff_rows = _build_diff_rows(sc_a['results'], sc_b['results'], row_key_from_obj)

        def _sum_diff(line_type):
            rows = [row for row in diff_rows if row['line_type'] == line_type]
            first = sc_a['results'].get(line_type, [])
            periods = list(first[0].get('values', {}).keys()) if first else []
            return {period: round(sum(row['diff'].get(period, 0) for row in rows), 2) for period in periods}

        summary = {
            'scenario_a_name': sc_a['name'],
            'scenario_b_name': sc_b['name'],
            'total_demand_diff': _sum_diff('03. Total demand'),
            'inventory_diff': _sum_diff('04. Inventory'),
            'changed_rows': len(diff_rows),
        }
        return jsonify({'summary': summary, 'rows': diff_rows})

    @bp.route('/api/scenarios/compare/export')
    def export_scenario_comparison():
        id_a = request.args.get('a', '')
        id_b = request.args.get('b', '')
        if id_a not in scenarios or id_b not in scenarios:
            return jsonify({'error': 'Scenario not found'}), 404
        sc_a = scenarios[id_a]
        sc_b = scenarios[id_b]
        if sc_a['session_id'] != get_active_session_id() or sc_b['session_id'] != get_active_session_id():
            return jsonify({'error': 'Scenarios belong to a different session'}), 403

        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
        row_a_fill = PatternFill(start_color='EBF3FB', end_color='EBF3FB', fill_type='solid')
        row_b_fill = PatternFill(start_color='FEF9EE', end_color='FEF9EE', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        bold_font = Font(bold=True)

        def _write_sheet(ws, diff_rows):
            if not diff_rows:
                ws.append(['No differences found.'])
                return
            periods = list(diff_rows[0]['values_a'].keys())
            headers = ['Material Number', 'Material Name', 'Line Type', 'Row'] + periods + ['Total Diff']
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for diff_row in diff_rows:
                mat = diff_row['material_number']
                name = diff_row['material_name']
                line_type = diff_row['line_type']
                values_a = diff_row['values_a']
                values_b = diff_row['values_b']
                diff = diff_row['diff']

                row_a_data = [mat, name, line_type, sc_a['name']]
                row_a_data += [values_a.get(period, 0) for period in periods]
                row_a_data += [round(sum(values_a.get(period, 0) for period in periods), 2)]
                ws.append(row_a_data)
                for cell in ws[ws.max_row]:
                    cell.fill = row_a_fill

                row_b_data = [mat, name, line_type, sc_b['name']]
                row_b_data += [values_b.get(period, 0) for period in periods]
                row_b_data += [round(sum(values_b.get(period, 0) for period in periods), 2)]
                ws.append(row_b_data)
                for cell in ws[ws.max_row]:
                    cell.fill = row_b_fill

                total_diff = round(sum(diff.get(period, 0) for period in periods), 2)
                diff_data = [mat, name, line_type, 'Diff (A-B)']
                diff_data += [diff.get(period, 0) for period in periods]
                diff_data += [total_diff]
                ws.append(diff_data)
                diff_excel_row = ws.max_row
                for col_idx, period in enumerate(periods, start=5):
                    cell = ws.cell(row=diff_excel_row, column=col_idx)
                    cell.font = bold_font
                    value = diff.get(period, 0)
                    if value > 0.01:
                        cell.fill = green_fill
                    elif value < -0.01:
                        cell.fill = red_fill

                total_cell = ws.cell(row=diff_excel_row, column=len(headers))
                total_cell.font = bold_font
                if total_diff > 0.01:
                    total_cell.fill = green_fill
                elif total_diff < -0.01:
                    total_cell.fill = red_fill

                ws.append([''] * len(headers))
                for cell in ws[ws.max_row]:
                    cell.fill = grey_fill

            for col_idx in range(1, 5):
                max_len = max(
                    (len(str(ws.cell(row_idx, col_idx).value or '')) for row_idx in range(1, ws.max_row + 1)),
                    default=10,
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

        wb = openpyxl.Workbook()
        ws_vol = wb.active
        ws_vol.title = 'Volume Comparison'
        vol_diff = _build_diff_rows(sc_a['results'], sc_b['results'], row_key_from_obj, include_name=True)
        _write_sheet(ws_vol, vol_diff)

        ws_val = wb.create_sheet('Value Comparison')
        val_diff = _build_diff_rows(
            sc_a.get('value_results', {}),
            sc_b.get('value_results', {}),
            row_key_from_obj,
            include_name=True,
        )
        _write_sheet(ws_val, val_diff)

        out_dir = export_dir()
        out_dir.mkdir(exist_ok=True)
        safe_a = ''.join(c for c in sc_a['name'] if c.isalnum() or c in ' _-')[:30]
        safe_b = ''.join(c for c in sc_b['name'] if c.isalnum() or c in ' _-')[:30]
        filename = f'Comparison_{safe_a}_vs_{safe_b}.xlsx'
        export_path = out_dir / filename
        wb.save(str(export_path))
        return send_file(str(export_path), as_attachment=True, download_name=filename)

    return bp


def _build_diff_rows(res_a, res_b, row_key_from_obj, include_name=False):
    rows = []
    for line_type, rows_a in res_a.items():
        rows_b_map = defaultdict(list)
        for row in res_b.get(line_type, []):
            rows_b_map[row_key_from_obj(row)].append(row)
        for row_a in rows_a:
            key = row_key_from_obj(row_a)
            bucket = rows_b_map.get(key, [])
            row_b = bucket.pop(0) if bucket else None
            if not row_b:
                continue
            diff = {
                period: round(row_a['values'].get(period, 0) - row_b['values'].get(period, 0), 4)
                for period in row_a['values']
            }
            if any(abs(value) > 0.01 for value in diff.values()):
                row = {
                    'material_number': row_a['material_number'],
                    'line_type': line_type,
                    'values_a': row_a['values'],
                    'values_b': row_b['values'],
                    'diff': diff,
                }
                if include_name:
                    row['material_name'] = row_a.get('material_name', '')
                rows.append(row)
    return rows
