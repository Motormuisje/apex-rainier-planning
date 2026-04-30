"""S&OP Planning Engine - Flask Web UI"""

from flask import Flask, request, jsonify
from pathlib import Path
import sys
import io
import os
import contextlib

from ui.parsers import (
    format_purchased_and_produced as _format_purchased_and_produced,
    parse_purchased_and_produced as _parse_purchased_and_produced,
    valuation_params_from_config as _valuation_params_from_config,
)
from ui.paths import default_app_data_root, default_folders, resource_root
from ui.serializers import (
    moq_warnings_payload as _moq_warnings_payload,
    planning_value_payload as _planning_value_payload,
    row_payload as _row_payload,
    value_results_payload as _value_results_payload,
)
from ui.config_store import (
    apply_folder_config,
    load_global_config,
    save_global_config,
    sync_global_config_from_engine,
)
from ui.cascade import (
    finish_pap_recalc as _finish_pap_recalc_impl,
    recalc_pap_material as _recalc_pap_material_impl,
)
from ui.errors import classify_upload_exception as _classify_upload_exception
from ui.replay import (
    recalculate_value_results,
    replay_pending_edits,
)
from ui.routes.config import create_config_blueprint
from ui.routes.edit_state import create_edit_state_blueprint
from ui.routes.edits import create_edits_blueprint
from ui.routes.exports import create_exports_blueprint
from ui.routes.license import create_license_blueprint
from ui.routes.machines import create_machines_blueprint
from ui.routes.pap import create_pap_blueprint
from ui.routes.read import create_read_blueprint
from ui.routes.scenarios import create_scenarios_blueprint
from ui.routes.sessions import create_sessions_blueprint
from ui.routes.workflow import create_workflow_blueprint
from ui.engine_rebuild import (
    build_clean_engine_for_session,
    get_config_overrides,
    get_session_config_overrides,
    install_clean_engine_baseline,
)
from ui.session_store import (
    load_sessions_from_disk,
    save_sessions_to_disk,
)
from ui.state_snapshot import (
    apply_machine_overrides as _apply_machine_overrides,
    build_pending_edits_from_results_snapshot as _build_pending_edits_from_results_snapshot,
    engine_has_manual_edits as _engine_has_manual_edits,
    ensure_reset_baseline,
    machine_overrides_from_engine as _machine_overrides_from_engine,
    planning_row_from_snapshot as _planning_row_from_snapshot,
    rebuild_volume_caches_from_results as _rebuild_volume_caches_from_results,
    restore_engine_state,
    row_key_from_obj as _row_key_from_obj,
    snapshot_has_manual_edits as _snapshot_has_manual_edits,
)
from ui.volume_change import (
    EDITABLE_LINE_TYPES,
    VALUE_AUX_EDITABLE_LINE_TYPES,
    SHIFT_HOURS_LOOKUP_FALLBACK,
    apply_volume_change,
    recalculate_capacity_and_values,
    recalc_one_material,
)

RESOURCE_ROOT = resource_root()
APP_DATA_ROOT = default_app_data_root()
APP_DATA_ROOT.mkdir(parents=True, exist_ok=True)

# Default folder paths â€” overridden by _apply_folder_config() after config loads
APP_UPLOADS_DIR = APP_DATA_ROOT / 'uploads'
APP_EXPORTS_DIR = APP_DATA_ROOT / 'exports'
APP_UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
APP_EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

def _default_folders() -> dict:
    return default_folders(APP_DATA_ROOT)

sys.path.insert(0, str(RESOURCE_ROOT))

from modules.planning_engine import PlanningEngine
from modules.models import LineType
from modules.cycle_manager import CycleManager
from modules.license_manager import LicenseManager

_license = LicenseManager(APP_DATA_ROOT)

app = Flask(
    __name__,
    template_folder=str(RESOURCE_ROOT / 'ui' / 'templates'),
    static_folder=str(RESOURCE_ROOT / 'ui' / 'static'),
)
# Allow large Excel uploads (512 MB). Without this, Flask returns a 413 that
# often surfaces on the client as "TypeError: Failed to fetch".
app.config['MAX_CONTENT_LENGTH'] = 512 * 1024 * 1024
app.register_blueprint(create_license_blueprint(_license))


@app.errorhandler(413)
def _too_large(e):
    return jsonify({
        'error': 'Bestand is te groot voor upload (limiet 512 MB). Comprimeer of splits het bestand.',
        'error_kind': 'too_large',
    }), 413

sessions: dict = {}           # session_id -> session dict
active_session_id: str = None  # currently selected session
scenarios: dict = {}          # scenario_id -> scenario snapshot

# Shared CycleManager â€” stores previous-cycle snapshots in the writable app-data exports folder
_CYCLE_STORAGE_DIR = APP_EXPORTS_DIR
_cycle_manager = CycleManager(str(_CYCLE_STORAGE_DIR))


SESSIONS_STORE = APP_DATA_ROOT / 'sessions_store.json'
GLOBAL_CONFIG_FILE = APP_DATA_ROOT / 'global_config.json'

_global_config: dict = {}
_VERBOSE_STARTUP = os.getenv('SOP_VERBOSE_STARTUP', '').strip().lower() in ('1', 'true', 'yes', 'on')
_DISABLE_AUTORUN = os.getenv('SOP_DISABLE_AUTORUN', '').strip().lower() in ('1', 'true', 'yes', 'on')


def _restore_engine_state(engine, snapshot: dict) -> None:
    restore_engine_state(engine, snapshot, _global_config)


def _install_clean_engine_baseline(sess, engine, clear_machine_overrides: bool = True) -> None:
    install_clean_engine_baseline(
        sess,
        engine,
        SHIFT_HOURS_LOOKUP_FALLBACK,
        clear_machine_overrides=clear_machine_overrides,
    )


def _load_global_config():
    global _global_config
    _global_config = load_global_config(GLOBAL_CONFIG_FILE)


def _save_global_config():
    save_global_config(GLOBAL_CONFIG_FILE, _global_config)


def _get_config_overrides() -> dict:
    return get_config_overrides(_global_config)


def _save_sessions_to_disk():
    try:
        save_sessions_to_disk(
            sessions,
            active_session_id,
            SESSIONS_STORE,
            _machine_overrides_from_engine,
        )
    except Exception as exc:
        print(f'[sessions] save error: {exc}')


def _load_sessions_from_disk():
    global sessions, active_session_id
    sessions, active_session_id = load_sessions_from_disk(SESSIONS_STORE)


def _apply_folder_config():
    """Apply folder paths from _global_config, update globals and CycleManager."""
    global APP_UPLOADS_DIR, APP_EXPORTS_DIR, SESSIONS_STORE, _cycle_manager
    APP_UPLOADS_DIR, APP_EXPORTS_DIR, SESSIONS_STORE = apply_folder_config(
        _global_config,
        _default_folders(),
    )
    _cycle_manager = CycleManager(str(APP_EXPORTS_DIR))


def _apply_folder_paths(uploads_dir: Path, exports_dir: Path, sessions_dir: Path) -> None:
    global APP_UPLOADS_DIR, APP_EXPORTS_DIR, SESSIONS_STORE, _cycle_manager
    APP_UPLOADS_DIR = uploads_dir
    APP_EXPORTS_DIR = exports_dir
    SESSIONS_STORE = sessions_dir / 'sessions_store.json'
    _cycle_manager = CycleManager(str(APP_EXPORTS_DIR))


_load_global_config()
_apply_folder_config()
_load_sessions_from_disk()
app.register_blueprint(create_config_blueprint(
    _default_folders,
    _global_config,
    _save_global_config,
    _apply_folder_paths,
    lambda: APP_UPLOADS_DIR,
    lambda: _get_active(),
    _parse_purchased_and_produced,
    _valuation_params_from_config,
    lambda sess, engine: ensure_reset_baseline(sess, engine, SHIFT_HOURS_LOOKUP_FALLBACK),
    lambda engine, material_number: _recalc_pap_material(engine, material_number),
    lambda engine: _finish_pap_recalc(engine),
    lambda engine, sess=None: recalculate_value_results(engine, sess),
    lambda sess, params=None: build_clean_engine_for_session(sess, _global_config, params),
    _install_clean_engine_baseline,
    lambda sess, engine: _replay_pending_edits(sess, engine),
    _moq_warnings_payload,
    _value_results_payload,
))
app.register_blueprint(create_read_blueprint(
    lambda: _get_active(),
    _row_payload,
    _moq_warnings_payload,
))
app.register_blueprint(create_machines_blueprint(
    lambda: _get_active(),
    _machine_overrides_from_engine,
    lambda machine, data: SHIFT_HOURS_LOOKUP_FALLBACK(machine, data),
    lambda sess, engine: ensure_reset_baseline(sess, engine, SHIFT_HOURS_LOOKUP_FALLBACK),
    lambda engine, sess: recalculate_capacity_and_values(engine, sess),
    _planning_value_payload,
    _save_sessions_to_disk,
))
app.register_blueprint(create_pap_blueprint(
    lambda: _get_active(),
    _global_config,
    _format_purchased_and_produced,
    lambda sess, engine: ensure_reset_baseline(sess, engine, SHIFT_HOURS_LOOKUP_FALLBACK),
    lambda engine, material_number: _recalc_pap_material(engine, material_number),
    lambda engine: _finish_pap_recalc(engine),
    _save_global_config,
    _moq_warnings_payload,
))
app.register_blueprint(create_sessions_blueprint(
    sessions,
    lambda: active_session_id,
    lambda session_id: _set_active_session_id(session_id),
    lambda: _get_active(),
    _global_config,
    _machine_overrides_from_engine,
    _save_sessions_to_disk,
    lambda engine: sync_global_config_from_engine(engine, _global_config, _format_purchased_and_produced),
    lambda sess, params=None: build_clean_engine_for_session(sess, _global_config, params),
    _install_clean_engine_baseline,
    lambda sess, engine: _replay_pending_edits(sess, engine),
    _snapshot_has_manual_edits,
    _engine_has_manual_edits,
    lambda: app.app_context(),
))
app.register_blueprint(create_scenarios_blueprint(
    scenarios,
    sessions,
    lambda: active_session_id,
    lambda: _get_active(),
    _global_config,
    lambda: APP_EXPORTS_DIR,
    _build_pending_edits_from_results_snapshot,
    _planning_row_from_snapshot,
    _rebuild_volume_caches_from_results,
    _valuation_params_from_config,
    _parse_purchased_and_produced,
    _format_purchased_and_produced,
    _row_key_from_obj,
))
app.register_blueprint(create_exports_blueprint(
    lambda: _get_active(),
    lambda: APP_EXPORTS_DIR,
    lambda: _cycle_manager,
    lambda path, engine: _apply_edit_highlights(path, engine),
))
app.register_blueprint(create_edit_state_blueprint(
    sessions,
    EDITABLE_LINE_TYPES,
    _save_sessions_to_disk,
))
app.register_blueprint(create_edits_blueprint(
    lambda: _get_active(),
    VALUE_AUX_EDITABLE_LINE_TYPES,
    _global_config,
    lambda *args, **kwargs: apply_volume_change(*args, **kwargs),
    lambda sess, engine: ensure_reset_baseline(sess, engine, SHIFT_HOURS_LOOKUP_FALLBACK),
    lambda engine, sess: recalculate_value_results(engine, sess),
    _save_sessions_to_disk,
    _valuation_params_from_config,
    _restore_engine_state,
    _snapshot_has_manual_edits,
    lambda sess, params=None: build_clean_engine_for_session(sess, _global_config, params),
    _install_clean_engine_baseline,
))
app.register_blueprint(create_workflow_blueprint(
    sessions,
    lambda session_id: _set_active_session_id(session_id),
    lambda: _get_active(),
    lambda: APP_UPLOADS_DIR,
    _global_config,
    _classify_upload_exception,
    _get_config_overrides,
    lambda: _cycle_manager,
    _install_clean_engine_baseline,
    lambda sess, engine: _replay_pending_edits(sess, engine),
    _save_sessions_to_disk,
    lambda: app.app_context(),
))


def _replay_pending_edits(sess, engine):
    replay_pending_edits(
        sess,
        engine,
        apply_volume_change,
        _apply_machine_overrides,
        recalculate_capacity_and_values,
    )



def _autorun_sessions():
    """Background thread: re-execute the planning engine for every session that
    was previously run (identified by a non-None 'parameters' field).

    Runs silently and does not block server startup.
    """
    import threading

    def _worker():
        candidates = [
            (sid, sess) for sid, sess in sessions.items()
            if sess.get('parameters') is not None and (
                sess.get('extract_files') or Path(sess.get('file_path', '')).exists()
            )
        ]
        if not candidates:
            return
        for sid, sess in candidates:
            label = sess.get('custom_name') or sess.get('filename', sid)
            try:
                params = sess['parameters']
                planning_month  = params.get('planning_month')
                months_actuals  = int(params.get('months_actuals', 0) or 0)
                months_forecast = int(params.get('months_forecast', 12) or 12)
                engine = PlanningEngine(
                    sess['file_path'],
                    planning_month=planning_month,
                    months_actuals=months_actuals,
                    months_forecast=months_forecast,
                    extract_files=sess.get('extract_files'),
                    config_overrides=get_session_config_overrides(sess, _global_config),
                )
                if _VERBOSE_STARTUP:
                    engine.run()
                    _install_clean_engine_baseline(sess, engine, clear_machine_overrides=False)
                    with app.app_context():
                        _replay_pending_edits(sess, engine)
                else:
                    with contextlib.redirect_stdout(io.StringIO()):
                        engine.run()
                        _install_clean_engine_baseline(sess, engine, clear_machine_overrides=False)
                        with app.app_context():
                            _replay_pending_edits(sess, engine)
                sess['engine'] = engine
            except Exception as exc:
                import logging
                logging.getLogger(__name__).error(f'autorun FAIL "{label}": {exc}')

    t = threading.Thread(target=_worker, name='autorun-sessions', daemon=True)
    t.start()


if not _DISABLE_AUTORUN:
    _autorun_sessions()


def _get_active():
    """Return (session_dict, engine) for the active session, or (None, None)."""
    sess = sessions.get(active_session_id)
    if not sess:
        return None, None
    return sess, sess.get('engine')


def _set_active_session_id(session_id):
    global active_session_id
    active_session_id = session_id



def _apply_edit_highlights(path: str, engine):
    """Open the exported workbook and apply edit highlights + summary sheet."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.comments import Comment

    # Collect all edit
    all_edits = []
    for _, rows in engine.results.items():
        for row in rows:
            if row.manual_edits:
                for period, edit_data in row.manual_edits.items():
                    original = edit_data.get('original', 0.0)
                    new_val = edit_data.get('new', 0.0)
                    delta_pct = round((new_val - original) / abs(original) * 100, 2) if original != 0 else 0.0
                    all_edits.append({
                        'line_type': row.line_type,
                        'material_number': row.material_number,
                        'material_name': row.material_name,
                        'period': period,
                        'original': original,
                        'new': new_val,
                        'delta_pct': delta_pct,
                    })

    if not all_edits:
        return

    wb = openpyxl.load_workbook(path)
    ws = wb['Planning sheet']

    # Build column lookups from header row
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    period_col = {}
    mat_col_idx = None
    lt_col_idx = None
    for i, val in enumerate(header, start=1):
        if val is None:
            continue
        s = str(val)
        period_col[s] = i
        if s == 'Material number':
            mat_col_idx = i
        elif s == 'Line type':
            lt_col_idx = i

    # Build row lookup: (material_number, line_type) -> row_idx
    row_lookup = {}
    if mat_col_idx and lt_col_idx:
        for row_idx, row_data in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            mat_val = row_data[mat_col_idx - 1]
            lt_val = row_data[lt_col_idx - 1]
            if mat_val and lt_val:
                row_lookup[(str(mat_val), str(lt_val))] = row_idx

    # Fill styles
    yellow_fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')
    green_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    red_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
    bold_font = Font(bold=True)

    for edit in all_edits:
        row_idx = row_lookup.get((edit['material_number'], edit['line_type']))
        col_idx = period_col.get(edit['period'])
        if row_idx is None or col_idx is None:
            continue
        cell = ws.cell(row=row_idx, column=col_idx)
        original = edit['original']
        new_val = edit['new']
        delta_pct = edit['delta_pct']
        if new_val > original:
            cell.fill = green_fill
            cell.font = bold_font
        elif new_val < original:
            cell.fill = red_fill
            cell.font = bold_font
        else:
            cell.fill = yellow_fill
        cell.comment = Comment(f"Original: {original}\nNew: {new_val}\nDelta: {delta_pct}%", 'SOP Engine')

    # Edits Summary sheet
    if 'Edits Summary' in wb.sheetnames:
        del wb['Edits Summary']
    ws_edits = wb.create_sheet('Edits Summary')
    ws_edits.append(['Line Type', 'Material Number', 'Material Name', 'Period',
                     'Original Value', 'New Value', 'Delta %'])
    for edit in all_edits:
        ws_edits.append([edit['line_type'], edit['material_number'], edit['material_name'],
                         edit['period'], edit['original'], edit['new'], edit['delta_pct']])

    wb.save(path)


# ---- Prod/Purch Split endpoints ----

def _recalc_pap_material(current_engine, material_number):
    _recalc_pap_material_impl(current_engine, material_number, recalc_one_material)

def _finish_pap_recalc(current_engine):
    sess = sessions.get(active_session_id) if active_session_id else None
    _finish_pap_recalc_impl(current_engine, sess, recalculate_capacity_and_values)


_SESSION_SAVE_PATHS = {
    '/api/sessions/rename',
    '/api/sessions/switch',
    '/api/sessions/snapshot',
    '/api/sessions/edits/persist',
    '/api/sessions/edits/sync',
    '/api/upload',
    '/api/calculate',
    '/api/update_volume',
    '/api/undo',
    '/api/redo',
    '/api/reset_edits',
    '/api/reset_value_planning_edits',
    '/api/scenarios/save',
    '/api/scenarios/load',
    '/api/export_db',
}

_SESSION_SAVE_METHODS = {'POST', 'DELETE'}


@app.after_request
def _after_request_save(response):
    """Auto-save sessions to disk after any mutating request."""
    if request.method in _SESSION_SAVE_METHODS and (
        request.path in _SESSION_SAVE_PATHS
        or (request.method == 'DELETE' and request.path.startswith('/api/sessions/'))
        or (request.method == 'DELETE' and request.path.startswith('/api/scenarios/'))
    ):
        if response.status_code < 500:
            _save_sessions_to_disk()
    return response


if __name__ == '__main__':
    host = os.getenv('SOP_HOST', '127.0.0.1')
    port = int(os.getenv('SOP_PORT', '5000'))
    app.run(debug=False, host=host, port=port)



