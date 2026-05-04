from types import SimpleNamespace

import pandas as pd
import pytest
from flask import Flask

import ui.routes.exports as exports_module
from modules.models import LineType, PlanningRow
from ui.routes.exports import _apply_edit_highlights, create_exports_blueprint


pytestmark = pytest.mark.no_fixture


def _make_planning_workbook(path, *, rows):
    """
    Create a minimal 'Planning sheet' workbook at `path`.
    `rows` is a list of dicts with keys:
        material_number, line_type, period_values (dict[period, value])
    Returns the list of period column names in the order they were written.
    """
    import openpyxl

    all_periods = list(dict.fromkeys(
        period for row in rows for period in row["period_values"]
    ))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planning sheet"
    ws.append(["Material number", "Name", "Line type"] + all_periods)
    for row in rows:
        ws.append([
            row["material_number"],
            row.get("material_name", f"Name {row['material_number']}"),
            row["line_type"],
        ] + [row["period_values"].get(period, 0.0) for period in all_periods])
    wb.save(str(path))
    return all_periods


def _engine_with_edit(line_type, material_number, period, original, new_val):
    row = PlanningRow(
        material_number=material_number,
        material_name=f"Name {material_number}",
        product_type="Bulk Product",
        product_family="",
        spc_product="",
        product_cluster="",
        product_name="",
        line_type=line_type,
        values={period: new_val},
        manual_edits={period: {"original": original, "new": new_val}},
    )
    return SimpleNamespace(results={line_type: [row]})


def _load_highlighted_cell(path, period, row_idx=2):
    import openpyxl

    wb = openpyxl.load_workbook(str(path))
    ws = wb["Planning sheet"]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    period_col_idx = header.index(period) + 1
    return wb, ws.cell(row=row_idx, column=period_col_idx)


class FakeMoMEngine:
    def to_dataframe(self):
        return pd.DataFrame([
            {
                "Line type": "04. Inventory",
                "Material number": "MAT-1",
                "Material name": "Material 1",
                "Product type": "Bulk Product",
                "2025-12": 10,
                "2026-01": 12,
                "2026-02": 9,
            },
            {
                "Line type": "04. Inventory",
                "Material number": "MAT-2",
                "Material name": "Material 2",
                "Product type": "Raw Material",
                "2025-12": 5,
                "2026-01": 7,
                "2026-02": 8,
            },
        ])


class FakeCycleManager:
    def __init__(self, previous_df=None):
        self.previous_df = previous_df

    def has_previous_cycle(self):
        return self.previous_df is not None

    def load_previous_cycle(self):
        return self.previous_df


class FakeExportEngine(FakeMoMEngine):
    def __init__(self):
        self.data = SimpleNamespace(config=SimpleNamespace(site="NLX1", initial_date="2025-12-01"))
        self.results = {}
        self.value_results = {}
        self.excel_calls = []

    def to_excel_with_values(self, path, inventory_quality_engine=None, previous_cycle_df=None):
        self.excel_calls.append({
            "path": path,
            "inventory_quality_engine": inventory_quality_engine,
            "previous_cycle_df": previous_cycle_df,
        })
        with open(path, "wb") as handle:
            handle.write(b"fake planning workbook")


@pytest.fixture
def exports_route_app(tmp_path, monkeypatch):
    active = {"engine": None}
    state = {
        "cycle_manager": FakeCycleManager(),
        "highlights": [],
    }

    def get_active():
        return {"id": "exports-session"}, active["engine"]

    monkeypatch.setattr(
        exports_module,
        "_apply_edit_highlights",
        lambda path, engine: state["highlights"].append((path, engine)),
    )

    flask_app = Flask(__name__)
    flask_app.config["TESTING"] = True
    flask_app.register_blueprint(create_exports_blueprint(
        get_active,
        lambda: tmp_path / "exports",
        lambda: state["cycle_manager"],
    ))

    return SimpleNamespace(
        app=flask_app,
        client=flask_app.test_client(),
        active=active,
        state=state,
        export_dir=tmp_path / "exports",
    )


def test_mom_returns_unavailable_without_engine(exports_route_app):
    response = exports_route_app.client.get("/api/mom")

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["available"] is False
    assert "Run calculations first" in payload["message"]


def test_mom_returns_sequential_comparison_from_dataframe(exports_route_app):
    exports_route_app.active["engine"] = FakeMoMEngine()

    response = exports_route_app.client.get("/api/mom?num_months=2")

    assert response.status_code == 200, response.get_json(silent=True) or response.get_data(as_text=True)
    payload = response.get_json()
    assert payload["available"] is True
    assert payload["periods"] == ["2025-12", "2026-01", "2026-02"]
    assert payload["material_count"] == 2
    assert payload["num_transitions"] == 2
    assert set(payload["scatter"]) == {"materials", "start", "end", "colors"}
    assert len(payload["summary"]) == 2
    assert len(payload["transitions"]) == 2


def test_export_requires_engine(exports_route_app):
    response = exports_route_app.client.get("/api/export")

    assert response.status_code == 400
    assert response.get_json() == {"error": "No calculations run"}


def test_export_writes_workbook_and_applies_highlights(exports_route_app):
    previous_df = pd.DataFrame([{"Line type": "04. Inventory", "2025-12": 1}])
    engine = FakeExportEngine()
    exports_route_app.active["engine"] = engine
    exports_route_app.state["cycle_manager"] = FakeCycleManager(previous_df)

    response = exports_route_app.client.get("/api/export")

    assert response.status_code == 200, response.get_json(silent=True) or response.get_data(as_text=True)
    assert response.headers["Content-Disposition"].startswith("attachment;")
    assert len(engine.excel_calls) == 1
    assert engine.excel_calls[0]["previous_cycle_df"].equals(previous_df)
    export_path = engine.excel_calls[0]["path"]
    assert export_path.endswith(".xlsx")
    assert exports_route_app.state["highlights"] == [(export_path, engine)]


def test_export_db_requires_engine(exports_route_app):
    response = exports_route_app.client.post("/api/export_db")

    assert response.status_code == 400
    assert response.get_json() == {"error": "No calculations run"}


def test_export_db_returns_400_when_exporter_has_no_rows(exports_route_app, monkeypatch):
    class EmptyExporter:
        def __init__(self, planning_df, site, initial_date):
            self.planning_df = planning_df
            self.site = site
            self.initial_date = initial_date

        def export_to_dataframe(self):
            return pd.DataFrame()

    exports_route_app.active["engine"] = FakeExportEngine()
    monkeypatch.setattr(exports_module, "DatabaseExporter", EmptyExporter)

    response = exports_route_app.client.post("/api/export_db")

    assert response.status_code == 400
    assert response.get_json() == {"error": "No data to export (no matching line types)"}


def test_export_db_writes_sanitized_filename(exports_route_app, monkeypatch):
    class OneRowExporter:
        def __init__(self, planning_df, site, initial_date):
            self.planning_df = planning_df
            self.site = site
            self.initial_date = initial_date

        def export_to_dataframe(self):
            return pd.DataFrame([{
                "site": self.site,
                "initial_date": self.initial_date,
                "rows": len(self.planning_df),
            }])

    exports_route_app.active["engine"] = FakeExportEngine()
    monkeypatch.setattr(exports_module, "DatabaseExporter", OneRowExporter)

    response = exports_route_app.client.post(
        "/api/export_db",
        json={"filename": "DB Export?!"},
    )

    assert response.status_code == 200, response.get_json(silent=True) or response.get_data(as_text=True)
    assert response.headers["Content-Disposition"] == "attachment; filename=\"DB Export.xlsx\""
    assert (exports_route_app.export_dir / "DB Export.xlsx").exists()


def test_apply_edit_highlights_no_edits_does_not_modify_file(tmp_path):
    path = tmp_path / "planning.xlsx"
    _make_planning_workbook(
        path,
        rows=[{
            "material_number": "MAT-1",
            "line_type": LineType.DEMAND_FORECAST.value,
            "period_values": {"2025-12": 100.0},
        }],
    )

    _apply_edit_highlights(str(path), SimpleNamespace(results={}))

    import openpyxl
    wb = openpyxl.load_workbook(str(path))
    assert "Edits Summary" not in wb.sheetnames


def test_apply_edit_highlights_increase_applies_green_fill(tmp_path):
    path = tmp_path / "planning.xlsx"
    _make_planning_workbook(
        path,
        rows=[{
            "material_number": "MAT-1",
            "line_type": LineType.DEMAND_FORECAST.value,
            "period_values": {"2025-12": 100.0},
        }],
    )
    engine = _engine_with_edit(LineType.DEMAND_FORECAST.value, "MAT-1", "2025-12", 90.0, 100.0)

    _apply_edit_highlights(str(path), engine)

    _, cell = _load_highlighted_cell(path, "2025-12")
    assert cell.fill.fgColor.rgb == "00C8E6C9"
    assert cell.font.bold is True
    assert cell.comment is not None
    assert "Original: 90.0" in cell.comment.text
    assert "New: 100.0" in cell.comment.text
    assert "Delta: 11.11%" in cell.comment.text
    assert cell.comment.author == "SOP Engine"


def test_apply_edit_highlights_decrease_applies_red_fill(tmp_path):
    path = tmp_path / "planning.xlsx"
    _make_planning_workbook(
        path,
        rows=[{
            "material_number": "MAT-1",
            "line_type": LineType.DEMAND_FORECAST.value,
            "period_values": {"2025-12": 80.0},
        }],
    )
    engine = _engine_with_edit(LineType.DEMAND_FORECAST.value, "MAT-1", "2025-12", 100.0, 80.0)

    _apply_edit_highlights(str(path), engine)

    _, cell = _load_highlighted_cell(path, "2025-12")
    assert cell.fill.fgColor.rgb == "00FFCDD2"
    assert cell.font.bold is True
    assert "Delta: -20.0%" in cell.comment.text


def test_apply_edit_highlights_zero_delta_applies_yellow_fill(tmp_path):
    path = tmp_path / "planning.xlsx"
    _make_planning_workbook(
        path,
        rows=[{
            "material_number": "MAT-1",
            "line_type": LineType.DEMAND_FORECAST.value,
            "period_values": {"2025-12": 100.0},
        }],
    )
    engine = _engine_with_edit(LineType.DEMAND_FORECAST.value, "MAT-1", "2025-12", 100.0, 100.0)

    _apply_edit_highlights(str(path), engine)

    _, cell = _load_highlighted_cell(path, "2025-12")
    assert cell.fill.fgColor.rgb == "00FFEB3B"
    assert cell.font.bold is False


def test_apply_edit_highlights_creates_edits_summary_sheet(tmp_path):
    path = tmp_path / "planning.xlsx"
    _make_planning_workbook(
        path,
        rows=[{
            "material_number": "MAT-1",
            "line_type": LineType.DEMAND_FORECAST.value,
            "period_values": {"2025-12": 100.0},
        }],
    )
    engine = _engine_with_edit(LineType.DEMAND_FORECAST.value, "MAT-1", "2025-12", 90.0, 100.0)

    _apply_edit_highlights(str(path), engine)

    import openpyxl
    wb = openpyxl.load_workbook(str(path))
    assert "Edits Summary" in wb.sheetnames
    ws_edits = wb["Edits Summary"]
    rows = list(ws_edits.iter_rows(values_only=True))
    assert rows[0] == (
        "Line Type", "Material Number", "Material Name", "Period",
        "Original Value", "New Value", "Delta %",
    )
    assert rows[1][0] == LineType.DEMAND_FORECAST.value
    assert rows[1][1] == "MAT-1"
    assert rows[1][3] == "2025-12"
    assert rows[1][4] == pytest.approx(90.0)
    assert rows[1][5] == pytest.approx(100.0)


def test_apply_edit_highlights_replaces_existing_summary_sheet(tmp_path):
    import openpyxl

    path = tmp_path / "planning.xlsx"
    _make_planning_workbook(
        path,
        rows=[{
            "material_number": "MAT-1",
            "line_type": LineType.DEMAND_FORECAST.value,
            "period_values": {"2025-12": 100.0},
        }],
    )
    wb = openpyxl.load_workbook(str(path))
    wb.create_sheet("Edits Summary").append(["stale"])
    wb.save(str(path))
    engine = _engine_with_edit(LineType.DEMAND_FORECAST.value, "MAT-1", "2025-12", 90.0, 100.0)

    _apply_edit_highlights(str(path), engine)

    wb = openpyxl.load_workbook(str(path))
    assert wb.sheetnames.count("Edits Summary") == 1
    rows = list(wb["Edits Summary"].iter_rows(values_only=True))
    assert rows[0][0] == "Line Type"
    assert all("stale" not in row for row in rows)
    assert rows[1][1] == "MAT-1"


def test_apply_edit_highlights_skips_edit_when_period_not_in_workbook(tmp_path):
    path = tmp_path / "planning.xlsx"
    _make_planning_workbook(
        path,
        rows=[{
            "material_number": "MAT-1",
            "line_type": LineType.DEMAND_FORECAST.value,
            "period_values": {"2025-12": 100.0},
        }],
    )
    engine = _engine_with_edit(LineType.DEMAND_FORECAST.value, "MAT-1", "2099-01", 90.0, 100.0)

    _apply_edit_highlights(str(path), engine)

    _, cell = _load_highlighted_cell(path, "2025-12")
    assert cell.fill.fgColor.rgb == "00000000"
    assert cell.comment is None

    import openpyxl
    wb = openpyxl.load_workbook(str(path))
    rows = list(wb["Edits Summary"].iter_rows(values_only=True))
    assert rows[1][1] == "MAT-1"
    assert rows[1][3] == "2099-01"


def test_apply_edit_highlights_skips_edit_when_material_not_in_workbook(tmp_path):
    path = tmp_path / "planning.xlsx"
    _make_planning_workbook(
        path,
        rows=[{
            "material_number": "MAT-1",
            "line_type": LineType.DEMAND_FORECAST.value,
            "period_values": {"2025-12": 100.0},
        }],
    )
    engine = _engine_with_edit(LineType.DEMAND_FORECAST.value, "MISSING", "2025-12", 90.0, 100.0)

    _apply_edit_highlights(str(path), engine)

    _, cell = _load_highlighted_cell(path, "2025-12")
    assert cell.fill.fgColor.rgb == "00000000"
    assert cell.comment is None

    import openpyxl
    wb = openpyxl.load_workbook(str(path))
    rows = list(wb["Edits Summary"].iter_rows(values_only=True))
    assert rows[1][1] == "MISSING"
    assert rows[1][3] == "2025-12"


def test_apply_edit_highlights_zero_original_uses_zero_delta(tmp_path):
    path = tmp_path / "planning.xlsx"
    _make_planning_workbook(
        path,
        rows=[{
            "material_number": "MAT-1",
            "line_type": LineType.DEMAND_FORECAST.value,
            "period_values": {"2025-12": 50.0},
        }],
    )
    engine = _engine_with_edit(LineType.DEMAND_FORECAST.value, "MAT-1", "2025-12", 0.0, 50.0)

    _apply_edit_highlights(str(path), engine)

    _, cell = _load_highlighted_cell(path, "2025-12")
    assert "Delta: 0.0%" in cell.comment.text


def test_apply_edit_highlights_multiple_edits_on_different_rows(tmp_path):
    path = tmp_path / "planning.xlsx"
    _make_planning_workbook(
        path,
        rows=[
            {
                "material_number": "MAT-1",
                "line_type": LineType.DEMAND_FORECAST.value,
                "period_values": {"2025-12": 100.0},
            },
            {
                "material_number": "MAT-2",
                "line_type": LineType.DEMAND_FORECAST.value,
                "period_values": {"2025-12": 80.0},
            },
        ],
    )
    row_1 = _engine_with_edit(LineType.DEMAND_FORECAST.value, "MAT-1", "2025-12", 90.0, 100.0).results[
        LineType.DEMAND_FORECAST.value
    ][0]
    row_2 = _engine_with_edit(LineType.DEMAND_FORECAST.value, "MAT-2", "2025-12", 100.0, 80.0).results[
        LineType.DEMAND_FORECAST.value
    ][0]
    engine = SimpleNamespace(results={LineType.DEMAND_FORECAST.value: [row_1, row_2]})

    _apply_edit_highlights(str(path), engine)

    wb, cell_1 = _load_highlighted_cell(path, "2025-12", row_idx=2)
    _, cell_2 = _load_highlighted_cell(path, "2025-12", row_idx=3)
    assert cell_1.fill.fgColor.rgb == "00C8E6C9"
    assert cell_2.fill.fgColor.rgb == "00FFCDD2"
    rows = list(wb["Edits Summary"].iter_rows(values_only=True))
    assert len(rows) == 3
    assert rows[1][1] == "MAT-1"
    assert rows[2][1] == "MAT-2"
