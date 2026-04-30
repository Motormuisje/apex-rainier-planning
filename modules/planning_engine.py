"""
S&OP Planning Engine - Main Orchestrator
Processes materials in BOM topological order (level by level).

For each BOM level:
  1. Aggregate dependent demand from parent levels
  2. Calculate total demand, target stock, production/purchase plan, inventory
  3. Compute dependent requirements -> feed to next level

Then: capacity utilization, shift availability, available capacity, utilization rate, FTE
"""

import pandas as pd
from typing import Dict, List, Optional
from datetime import datetime
from pathlib import Path
from collections import defaultdict

from modules.models import PlanningRow, LineType
from modules.data_loader import DataLoader
try:
    from modules.inventory_quality_engine import InventoryQualityEngine as _IQEngine
except ImportError:
    _IQEngine = None
from modules.forecast_engine import ForecastEngine
from modules.bom_engine import BOMEngine
from modules.inventory_engine import InventoryEngine
from modules.capacity_engine import CapacityEngine
from modules.value_planning_engine import ValuePlanningEngine


class PlanningEngine:
    """Main orchestrator that runs all planning calculations."""

    EXPECTED_LINE_TYPES = [
        LineType.DEMAND_FORECAST.value,
        LineType.DEPENDENT_DEMAND.value,
        LineType.TOTAL_DEMAND.value,
        LineType.INVENTORY.value,
        LineType.MIN_TARGET_STOCK.value,
        LineType.PRODUCTION_PLAN.value,
        LineType.PURCHASE_RECEIPT.value,
        LineType.PURCHASE_PLAN.value,
        LineType.CAPACITY_UTILIZATION.value,
        LineType.DEPENDENT_REQUIREMENTS.value,
        LineType.AVAILABLE_CAPACITY.value,
        LineType.UTILIZATION_RATE.value,
        LineType.SHIFT_AVAILABILITY.value,
        LineType.FTE_REQUIREMENTS.value,
        LineType.CONSOLIDATION.value,
    ]

    def __init__(self, file_path: str = None, planning_month: str = None,
                 months_actuals: int = 0, months_forecast: int = 12,
                 extract_files: dict = None, config_overrides: dict = None):
        self.file_path = file_path
        self.extract_files = extract_files
        self.planning_month = planning_month
        self.months_actuals = months_actuals
        self.months_forecast = months_forecast
        self.config_overrides = config_overrides or {}
        self.data: Optional[DataLoader] = None

        self.results: Dict[str, List[PlanningRow]] = {lt: [] for lt in self.EXPECTED_LINE_TYPES}
        self.all_rows: List[PlanningRow] = []
        self.summary: Dict = {}

        # Cross-material tracking
        self.all_production_plans: Dict[str, Dict[str, float]] = {}
        self.all_purchase_receipts: Dict[str, Dict[str, float]] = {}
        self.all_total_demands: Dict[str, Dict[str, float]] = {}
        self.all_purch_raw_needs: Dict[str, Dict[str, float]] = {}
        self.machine_throughput_theo: Dict[str, float] = {}
        self.output_by_machine_period: Dict[str, Dict[str, float]] = {}
        
        # Value planning (NEW)
        self.value_results: Dict[str, List[PlanningRow]] = {}
        self.value_engine: Optional[ValuePlanningEngine] = None

    def run(self) -> 'PlanningEngine':
        """Run the complete planning pipeline."""
        print("\n" + "=" * 70)
        print("S&OP PLANNING ENGINE - FULL CALCULATION")
        print("=" * 70)

        # ===== STEP 1: Load data =====
        print("\n[STEP 1] Loading raw input data...")
        if self.extract_files:
            self.data = DataLoader(excel_file=self.file_path, extract_files=self.extract_files,
                                   config_overrides=self.config_overrides)
        else:
            self.data = DataLoader(self.file_path, config_overrides=self.config_overrides)
        self.data.load_all()

        # ===== STEP 1b: Apply UI parameter overrides =====
        # All three UI values (planning_month, months_forecast, months_actuals) must be
        # applied here — before any engine is instantiated — so that data.periods and
        # data.forecast_actuals_months are the single source of truth for the full
        # pipeline.  Engines read these attributes directly; none of them re-read the
        # Excel Config sheet after this point.
        #
        # planning_month  → data.config.initial_date  → period window start
        #                                              → actuals/forecast boundary
        #                                              → opening inventory date
        # months_forecast → data.config.forecast_months → length of data.periods
        #                                              → every engine's period loop
        # months_actuals  → data.forecast_actuals_months → Aux2 start index in
        #                                              ForecastEngine._calculate_aux_columns

        _original_actuals = self.data.forecast_actuals_months
        if self.planning_month:
            # Copilot: VBA DefineVariables lines 3447-3449 anchor forecast columns
            # from InitialDate + ForecastActualsMonths, so preserve Config anchors first.
            _config_initial_date = self.data.config.initial_date
            # Copilot: VBA DefineVariables lines 3447-3449 use ForecastActualsMonths
            # to position ForecastStartClmn from ForecastActualStartClmn.
            _original_actuals = self.data.forecast_actuals_months
            _pm = None
            for _fmt in ('%Y/%m', '%Y-%m'):
                try:
                    _pm = datetime.strptime(self.planning_month, _fmt)
                    break
                except ValueError:
                    continue
            if _pm is None:
                print(f"  >> WARNING: Could not parse planning_month='{self.planning_month}' "
                      f"(expected YYYY/MM or YYYY-MM) — using date from Excel Config sheet")
            else:
                _config_initial_date = self.data.config.initial_date
                _original_actuals    = self.data.forecast_actuals_months
                self.data.config.initial_date = _pm
                _month_shift = (_pm.year  - _config_initial_date.year)  * 12 \
                             + (_pm.month - _config_initial_date.month)
                # VBA never modifies ForecastActualsMonths when InitialDate
                # changes — ForecastStartClmn is always anchored to the
                # original Config value.  Do NOT add _month_shift here.
                print(f"  >> planning_month shift: {_month_shift} months from Config initial_date")

        # Copilot: VBA DefineVariables line 3443 sets end column from ForecastMonths only:
        # PlanningEndForecastClmn = PlanningStartForecastClmn + ForecastMonths - 1
        total_horizon = self.months_forecast
        self.data.config.forecast_months = total_horizon
        print(f"  >> horizon override: {total_horizon} forecast periods")

        # Regenerate periods once — after both start-date and length are finalised.
        self.data.periods = self.data.config.get_periods()

        print(f"  >> Final horizon: {len(self.data.periods)} periods starting "
              f"{self.data.periods[0] if self.data.periods else 'N/A'}, "
              f"actuals={self.data.forecast_actuals_months}")

        # ===== STEP 2: Demand Forecast (Line 01) =====
        print("\n[STEP 2] Calculating Demand Forecast (Line 01)...")
        # VBA: ForecastStartClmn = ForecastActualStartClmn + ForecastActualsMonths + 1
        # ForecastActualsMonths is always the ORIGINAL Config value — never shifted.
        actuals_months = self.data.forecast_actuals_months
        forecast_months = self.months_forecast if self.months_forecast > 0 else self.data.config.forecast_months
        print(f"  >> USING: actuals_months={actuals_months}, forecast_months={forecast_months}")
        forecast_engine = ForecastEngine(self.data, actuals_months, forecast_months)
        forecast_rows = forecast_engine.calculate()
        self.results[LineType.DEMAND_FORECAST.value] = forecast_rows
        forecasts = forecast_engine.get_all_forecasts()

        # ===== STEP 3: BOM structure analysis =====
        print("\n[STEP 3] Analyzing BOM structure...")
        bom_engine = BOMEngine(self.data)

        # ===== STEP 4: Level-by-level calculation =====
        print("\n[STEP 4] Level-by-level planning calculation...")
        inv_engine = InventoryEngine(self.data)
        periods = self.data.periods

        # Accumulate dependent demand for each material from parents
        dep_demand_by_parent: Dict[str, Dict[str, Dict[str, float]]] = defaultdict(
            lambda: defaultdict(lambda: {p: 0.0 for p in periods})
        )
        dep_demand_agg: Dict[str, Dict[str, float]] = defaultdict(
            lambda: {p: 0.0 for p in periods}
        )

        max_level = bom_engine.get_max_level()

        for level in range(max_level + 1):
            materials_at_level = bom_engine.get_materials_at_level(level)
            print(f"\n  --- Level {level}: {len(materials_at_level)} materials ---")

            for mat_num in materials_at_level:
                material = self.data.materials.get(mat_num)
                if not material:
                    # VBA creates Line 02 from BOM data directly — no Material Master
                    # required. Still emit Line 02 for BOM children missing from MatMas.
                    mat_dep_by_parent = dict(dep_demand_by_parent.get(mat_num, {}))
                    if mat_dep_by_parent:
                        dd_rows = bom_engine.create_dependent_demand_rows(
                            mat_num, mat_dep_by_parent
                        )
                        self.results[LineType.DEPENDENT_DEMAND.value].extend(dd_rows)
                    continue

                if not material.is_active:
                    continue

                mat_forecast = forecasts.get(mat_num, {})
                mat_dep_agg = dict(dep_demand_agg.get(mat_num, {p: 0.0 for p in periods}))
                mat_dep_by_parent = dict(dep_demand_by_parent.get(mat_num, {}))

                # Create dependent demand rows (Line 02)
                if mat_dep_by_parent:
                    dd_rows = bom_engine.create_dependent_demand_rows(
                        mat_num, mat_dep_by_parent
                    )
                    self.results[LineType.DEPENDENT_DEMAND.value].extend(dd_rows)

                # Calculate inventory-related lines for this material
                result = inv_engine.calculate_for_material(
                    mat_num, mat_forecast, mat_dep_agg, mat_dep_by_parent
                )

                for row in result['rows']:
                    lt = row.line_type
                    if lt in self.results:
                        self.results[lt].append(row)

                if result['production_plan'] is not None:
                    self.all_production_plans[mat_num] = result['production_plan']
                if result['purchase_receipt'] is not None:
                    self.all_purchase_receipts[mat_num] = result['purchase_receipt']
                if result['total_demand']:
                    self.all_total_demands[mat_num] = result['total_demand']
                if result.get('purch_raw_need'):
                    self.all_purch_raw_needs[mat_num] = result['purch_raw_need']

                # ===== Compute dependent requirements (Line 08) =====
                # Always generate dep requirements if material has BOM children
                # (even if production plan is all zeros - Excel does this)
                prod_plan = result['production_plan']
                if prod_plan is not None:
                    children_demand = bom_engine.compute_dependent_requirements(
                        mat_num, prod_plan
                    )

                    if children_demand:
                        dr_rows = bom_engine.create_dependent_requirements_rows(
                            mat_num, children_demand
                        )
                        self.results[LineType.DEPENDENT_REQUIREMENTS.value].extend(dr_rows)

                        for child, child_period_data in children_demand.items():
                            for period in periods:
                                dep_demand_agg[child][period] += child_period_data.get(period, 0.0)
                            dep_demand_by_parent[child][mat_num] = child_period_data

        # ===== STEP 4b: Process standalone safety stock materials =====
        bom_processed = set()
        for level in range(max_level + 1):
            bom_processed.update(bom_engine.get_materials_at_level(level))

        standalone_mats = [
            mat_num for mat_num in self.data.safety_stock
            if mat_num not in bom_processed
            and mat_num in self.data.materials
            and self.data.materials[mat_num].is_active
        ]

        if standalone_mats:
            print(f"\n  --- Standalone safety stock materials: {len(standalone_mats)} ---")
            for mat_num in standalone_mats:
                mat_forecast = forecasts.get(mat_num, {})
                empty_demand = {p: 0.0 for p in periods}

                result = inv_engine.calculate_for_material(
                    mat_num, mat_forecast, empty_demand, {}
                )

                for row in result['rows']:
                    lt = row.line_type
                    if lt in self.results:
                        self.results[lt].append(row)

                if result['production_plan'] is not None:
                    self.all_production_plans[mat_num] = result['production_plan']
                if result['purchase_receipt'] is not None:
                    self.all_purchase_receipts[mat_num] = result['purchase_receipt']
                if result['total_demand']:
                    self.all_total_demands[mat_num] = result['total_demand']
                if result.get('purch_raw_need'):
                    self.all_purch_raw_needs[mat_num] = result['purch_raw_need']

        # ===== STEP 5: Capacity calculations =====
        # NLI1 Exception 2: B15 production plan adjustment (VBA Exceptions_GTB lines 1873+)
        # Must run AFTER all materials processed, BEFORE CapacityEngine.
        if self.data.config and getattr(self.data.config, 'site', None) == 'NLI1':
            B15   = '600004811'
            B4010 = '600004831'
            b4010_prod = self.all_production_plans.get(B4010, {})
            b15_prod   = self.all_production_plans.get(B15)
            if b15_prod is not None and b4010_prod:
                # Subtract B4010 production from B15 production plan (floor at 0)
                for p in self.data.periods:
                    b15_prod[p] = max(0.0, b15_prod[p] - b4010_prod.get(p, 0.0))
                # Update the Line 06 Production plan row for B15
                for row in self.results.get(LineType.PRODUCTION_PLAN.value, []):
                    if row.material_number == B15:
                        row.values = dict(b15_prod)
                        break
                # Recalculate B15 inventory: B4010 production counts as additional supply
                b15_inv_rows = [r for r in self.results.get(LineType.INVENTORY.value, [])
                                if r.material_number == B15]
                if b15_inv_rows:
                    b15_inv    = b15_inv_rows[0]
                    b15_demand = self.all_total_demands.get(B15, {})
                    b15_purch  = self.all_purchase_receipts.get(B15) or {}
                    running    = b15_inv.starting_stock
                    for p in self.data.periods:
                        demand           = b15_demand.get(p, 0.0)
                        prod             = b15_prod.get(p, 0.0)
                        purch            = b15_purch.get(p, 0.0)
                        b4010_contrib    = b4010_prod.get(p, 0.0)
                        running          = running - demand + prod + purch + b4010_contrib
                        b15_inv.values[p] = running
                print(f"[NLI1] Exception 2 applied: B15 production adjusted by B4010 contribution.")

        print("\n[STEP 5] Calculating capacity...")
        # Build per-line-type lookup: {line_type: {mat_num: {period: value}}}
        # VBA TruckOperationsFormulas SUMIFS references the truck row's own col C
        # (product_type_raw) as the line-type filter and col B (material_name) as
        # the product-type filter — so CapacityEngine needs all line-type data.
        all_line_data = {
            lt: {r.material_number: r.values for r in rows}
            for lt, rows in self.results.items()
            if rows
        }
        capacity_engine = CapacityEngine(self.data, self.all_production_plans, all_line_data)
        capacity_results = capacity_engine.calculate()
        for line_type, rows in capacity_results.items():
            self.results[line_type] = rows
        self.rebuild_machine_output_caches()
        
        # ===== STEP 6: Value planning calculations =====
        print("\n[STEP 6] Calculating value planning...")
        self.value_engine = ValuePlanningEngine(self.data, self.results)
        self.value_results = self.value_engine.calculate()

        # ===== Compile and validate =====
        self._compile_all_rows()
        self._validate_output()
        self._generate_summary()

        print("\n" + "=" * 70)
        print("CALCULATION COMPLETE")
        print("=" * 70)
        self._print_summary()

        return self

    def rebuild_machine_output_caches(self) -> None:
        """Precompute per-machine throughput data used by the machines route."""
        if self.data is None:
            self.machine_throughput_theo = {}
            self.output_by_machine_period = {}
            return

        periods = self.data.periods

        theo_lists = {}
        for mat_num in list(self.data.materials.keys()):
            try:
                routings = self.data.get_all_routings(mat_num)
            except Exception:
                continue
            for routing in routings:
                wc = routing.work_center
                if routing.base_quantity > 0 and routing.standard_time > 0:
                    theo_lists.setdefault(wc, []).append(routing.base_quantity / routing.standard_time)
        self.machine_throughput_theo = {
            wc: sum(values) / len(values) if values else 0.0
            for wc, values in theo_lists.items()
        }

        output_by_machine_period = {
            mc: {period: 0.0 for period in periods}
            for mc in self.data.machines
        }
        for mat_num, plan_data in self.all_production_plans.items():
            try:
                routings = self.data.get_all_routings(mat_num)
            except Exception:
                continue
            for routing in routings:
                wc = routing.work_center
                if wc not in output_by_machine_period:
                    continue
                for period in periods:
                    qty = plan_data.get(period, 0.0)
                    if qty > 0:
                        output_by_machine_period[wc][period] += qty
        self.output_by_machine_period = output_by_machine_period

    def _compile_all_rows(self):
        self.all_rows = []
        for line_type in self.EXPECTED_LINE_TYPES:
            if line_type in self.results:
                self.all_rows.extend(self.results[line_type])

    def _validate_output(self):
        print("\n[VALIDATION] Checking output...")
        active_types = [lt for lt, rows in self.results.items() if len(rows) > 0]
        print(f"  Line types with data: {len(active_types)}/{len(self.EXPECTED_LINE_TYPES)}")

        for lt in self.EXPECTED_LINE_TYPES:
            count = len(self.results.get(lt, []))
            status = "V" if count > 0 else "X"
            print(f"    {status} {lt}: {count} rows")

        if len(active_types) < 10:
            print(f"  WARNING: Only {len(active_types)} line types have data")
        else:
            print("  Validation passed")

    def _generate_summary(self):
        self.summary = {
            'total_rows': len(self.all_rows),
            'line_types_count': len([lt for lt, rows in self.results.items() if rows]),
            'line_types': {lt: len(rows) for lt, rows in self.results.items()},
            'materials': len(self.data.materials),
            'bom_items': len(self.data.bom),
            'machines': len(self.data.machines),
            'machine_groups': len(self.data.machine_groups),
            'periods': len(self.data.periods),
            'period_list': self.data.periods,
        }

    def _print_summary(self):
        print(f"\nSummary:")
        print(f"  Total rows: {self.summary['total_rows']}")
        print(f"  Line types: {self.summary['line_types_count']}")
        print(f"  Periods: {self.summary['periods']}")
        print(f"\nBreakdown by line type:")
        for lt in self.EXPECTED_LINE_TYPES:
            count = self.summary['line_types'].get(lt, 0)
            print(f"  {lt}: {count}")

    # ===== Export methods =====
    def get_all_rows(self) -> List[PlanningRow]:
        return self.all_rows

    def get_rows_by_type(self, line_type: str) -> List[PlanningRow]:
        return self.results.get(line_type, [])

    def get_summary(self) -> Dict:
        return self.summary

    def to_dataframe(self) -> pd.DataFrame:
        # Always rebuild from self.results to pick up cascade changes
        rows_data = []
        for line_type in self.EXPECTED_LINE_TYPES:
            for row in self.results.get(line_type, []):
                row_dict = {
                    'Material number': row.material_number,
                    'Material name': row.material_name,
                    'Product type': row.product_type,
                    'Product family': row.product_family,
                    'SPC product': row.spc_product,
                    'Product cluster': row.product_cluster,
                    'Product name': row.product_name,
                    'Line type': row.line_type,
                    'Aux Column': row.aux_column,
                    'Aux 2 Column': row.aux_2_column,
                    'Starting stock': row.starting_stock,
                }
                for period, value in row.values.items():
                    row_dict[period] = value
                rows_data.append(row_dict)
        df = pd.DataFrame(rows_data)
        # VBA SortPlanningSheet (line 4811): sort by material number ASC, line type ASC, aux1 ASC, aux2 ASC
        if not df.empty:
            # Cast sort-key columns to str so mixed types (str + numeric + None)
            # don't crash pandas rank().  Aux columns stay as-is (numeric) because
            # _sort_key handles them via pd.to_numeric.
            for _sc in ('Material number', 'Line type'):
                if _sc in df.columns:
                    df[_sc] = df[_sc].fillna('').astype(str)

            def _sort_key(col):
                if col.name in ('Aux Column', 'Aux 2 Column'):
                    return pd.to_numeric(col, errors='coerce').fillna(0)
                return col

            df = df.sort_values(
                by=['Material number', 'Line type', 'Aux Column', 'Aux 2 Column'],
                ascending=[True, True, True, True],
                na_position='last',
                key=_sort_key,
            ).reset_index(drop=True)
        return df

    def to_excel(self, output_path: str):
        df = self.to_dataframe()
        df.to_excel(output_path, sheet_name='Planning Results', index=False)
        print(f"\nResults exported to: {output_path}")
    
    def to_excel_with_values(self, output_path: str, inventory_quality_engine=None, previous_cycle_df=None):
        """Export both volume planning and value planning to Excel with VBA-matching formatting."""
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Volume planning sheet — written first and must stay visible so
            # openpyxl never sees an empty/all-hidden workbook on save.
            df_volumes = self.to_dataframe()
            df_volumes.to_excel(writer, sheet_name='Planning sheet', index=False)
            # Ensure the first sheet is never hidden (guards against IndexError
            # "At least one sheet must be visible" on partial-write failures).
            writer.book['Planning sheet'].sheet_state = 'visible'

            # Value planning sheet
            value_rows = []
            for line_type, rows in self.value_results.items():
                for row in rows:
                    row_dict = {
                        'Material number': row.material_number,
                        'Material name': row.material_name,
                        'Product type': row.product_type,
                        'Product family': row.product_family,
                        'SPC product': row.spc_product,
                        'Product cluster': row.product_cluster,
                        'Product name': row.product_name,
                        'Line type': row.line_type,
                        'Aux Column': row.aux_column,
                        'Aux 2 Column': row.aux_2_column,
                        'Starting stock': row.starting_stock,
                    }
                    for period, value in row.values.items():
                        row_dict[period] = value
                    value_rows.append(row_dict)

            df_values = pd.DataFrame(value_rows)
            df_values.to_excel(writer, sheet_name='Values_Planning sheet', index=False)

            # FTE requirements sheet (Line 12 rows)
            from modules.models import LineType
            fte_line = LineType.FTE_REQUIREMENTS.value
            fte_rows_data = self.results.get(fte_line, [])
            fte_records = []
            for row in fte_rows_data:
                rec = {
                    'Material number': row.material_number,
                    'Group name': row.material_name,
                    'FTE needed': row.aux_column,
                }
                for period, val in row.values.items():
                    rec[period] = round(val, 2)
                fte_records.append(rec)

            if fte_records:
                df_fte = pd.DataFrame(fte_records)
                period_cols = [c for c in df_fte.columns if str(c).count('-') == 1 and len(str(c)) == 7]

                # Add Average column
                if period_cols:
                    df_fte['Average'] = df_fte[period_cols].mean(axis=1).round(2)

                # Add TOTAL row
                total_rec = {'Material number': 'TOTAL', 'Group name': 'TOTAL', 'FTE needed': ''}
                for p in period_cols:
                    total_rec[p] = round(df_fte[p].sum(), 2)
                if period_cols:
                    total_rec['Average'] = round(df_fte['Average'].sum(), 2)
                df_fte = pd.concat([df_fte, pd.DataFrame([total_rec])], ignore_index=True)

                df_fte.to_excel(writer, sheet_name='FTE requirements', index=False)

                # Format FTE sheet
                ws_fte = writer.book['FTE requirements']
                self._apply_fte_formatting(ws_fte, period_cols)

            # Apply VBA-matching formatting to both sheets
            wb = writer.book
            self._apply_excel_formatting(wb['Planning sheet'])
            if not df_values.empty:
                self._apply_excel_formatting(wb['Values_Planning sheet'], is_values_sheet=True)
                self._apply_consol_colors(wb['Values_Planning sheet'])

            if 'Values_Planning sheet' in wb.sheetnames:
                pass  # no freeze panes

            # ---- High-level overview sheet (VBA CreateHighLevelOverview line 2804) ----
            from openpyxl.chart import BarChart, LineChart, Reference
            from openpyxl.chart.series import SeriesLabel

            ws_overview = wb.create_sheet('High-level overview')
            ws_overview.sheet_properties.tabColor = '8B4513'  # brown tab matching VBA

            # Collect consolidation rows
            consol_rows = []
            for lt, rows in self.value_results.items():
                for row in rows:
                    if row.line_type == LineType.CONSOLIDATION.value:
                        consol_rows.append(row)

            # Write header row: Metric | Summary | period1 | period2 | ...
            ws_overview['A1'] = 'Metric'
            ws_overview['B1'] = 'Summary'
            for i, p in enumerate(self.data.periods):
                ws_overview.cell(row=1, column=3 + i, value=p)

            # Write consolidation data rows
            for r_idx, row in enumerate(consol_rows, start=2):
                ws_overview.cell(row=r_idx, column=1, value=row.material_number.replace('ZZZZZZ_', ''))
                try:
                    ws_overview.cell(row=r_idx, column=2, value=float(row.aux_2_column) if row.aux_2_column else 0)
                except (ValueError, TypeError):
                    ws_overview.cell(row=r_idx, column=2, value=0)
                for i, p in enumerate(self.data.periods):
                    ws_overview.cell(row=r_idx, column=3 + i, value=row.get_value(p))

            n_periods = len(self.data.periods)
            n_consol = len(consol_rows)
            cats = Reference(ws_overview, min_col=3, max_col=2 + n_periods, min_row=1)

            # KPI summary table (VBA lines 2868-2940: 13 text box KPIs → formatted cells)
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            _ks = Side(style='thin')
            _kb = Border(left=_ks, right=_ks, top=_ks, bottom=_ks)
            _kh_fill = PatternFill(patternType='solid', fgColor='1F3864')   # dark navy header
            _kd_fill = PatternFill(patternType='solid', fgColor='D9E1F2')   # light blue data rows
            kpi_header_row = n_consol + 2
            kpi_items = consol_rows[:13]
            for _ci, _lbl in [(1, 'KPI'), (2, 'Summary (avg)')]:
                _c = ws_overview.cell(row=kpi_header_row, column=_ci, value=_lbl)
                _c.font = Font(bold=True, color='FFFFFF')
                _c.fill = _kh_fill
                _c.border = _kb
                _c.alignment = Alignment(horizontal='center')
            for _ki, _kitem in enumerate(kpi_items, start=1):
                _kr = kpi_header_row + _ki
                _klabel = _kitem.material_number.replace('ZZZZZZ_', '')
                try:
                    _kval = float(_kitem.aux_2_column) if _kitem.aux_2_column else 0.0
                except (ValueError, TypeError):
                    _kval = 0.0
                _lc = ws_overview.cell(row=_kr, column=1, value=_klabel)
                _lc.font = Font(bold=True)
                _lc.fill = _kd_fill
                _lc.border = _kb
                _vc = ws_overview.cell(row=_kr, column=2, value=round(_kval, 4))
                _vc.fill = _kd_fill
                _vc.border = _kb
                _vc.number_format = '#,##0.00'
                _vc.alignment = Alignment(horizontal='right')
            ws_overview.column_dimensions['A'].width = 26
            ws_overview.column_dimensions['B'].width = 18
            kpi_end_row = kpi_header_row + len(kpi_items)

            # Helper rows — ROCE target 15% and ROCE average constant
            target_row = kpi_end_row + 2
            roce_avg_row = kpi_end_row + 3
            ws_overview.cell(row=target_row, column=1, value='Target (15%)')
            for i in range(n_periods):
                ws_overview.cell(row=target_row, column=3 + i, value=0.15)
            _roce_consol = next(
                (r for r in consol_rows if 'ROCE' in r.material_number and 'CAPITAL' not in r.material_number),
                None
            )
            _roce_avg_val = 0.0
            if _roce_consol:
                try:
                    _roce_avg_val = float(_roce_consol.aux_2_column) if _roce_consol.aux_2_column else 0.0
                except (ValueError, TypeError):
                    _roce_avg_val = 0.0
            ws_overview.cell(row=roce_avg_row, column=1, value='ROCE average')
            for i in range(n_periods):
                ws_overview.cell(row=roce_avg_row, column=3 + i, value=_roce_avg_val)

            # Chart anchor base; charts 2 and 3 inserted in Phase 2 (after dedicated sheets exist)
            _chart_base = kpi_end_row + 5
            _ov_chart2_anchor = 'A' + str(_chart_base + 24)   # Inventory Quality (Phase 2)
            _ov_chart3_anchor = 'A' + str(_chart_base + 60)   # Top 10 Overstocks (Phase 2)

            # Chart 1 — Projected Financial Metrics (VBA step 141)
            chart1 = LineChart()
            chart1.title = 'Projected Financial Metrics'
            chart1.width = 20
            chart1.height = 12
            chart1.y_axis.title = 'Value'
            metrics_chart1 = ['TURNOVER', 'COST OF GOODS', 'GROSS MARGIN', 'INVENTORY VALUE']
            for metric in metrics_chart1:
                for r_idx, row in enumerate(consol_rows, start=2):
                    if row.material_number.replace('ZZZZZZ_', '') == metric:
                        data_ref = Reference(ws_overview, min_col=3, max_col=2 + n_periods, min_row=r_idx)
                        chart1.add_data(data_ref, titles_from_data=False)
                        chart1.series[-1].title = SeriesLabel(v=metric)
                        break
            chart1.set_categories(cats)
            ws_overview.add_chart(chart1, 'A' + str(_chart_base))

            # Chart 4 — ROCE Components (VBA step 144; 4th of 5 charts in sequence)
            chart2 = LineChart()
            chart2.title = 'ROCE Components'
            chart2.width = 20
            chart2.height = 12
            chart2.y_axis.title = 'Value'
            metrics_chart2 = ['EBIT', 'CAPITAL INVESTMENT', 'OPERATIONAL CASHFLOW']
            for metric in metrics_chart2:
                for r_idx, row in enumerate(consol_rows, start=2):
                    if row.material_number.replace('ZZZZZZ_', '') == metric:
                        data_ref = Reference(ws_overview, min_col=3, max_col=2 + n_periods, min_row=r_idx)
                        chart2.add_data(data_ref, titles_from_data=False)
                        chart2.series[-1].title = SeriesLabel(v=metric)
                        break
            chart2.set_categories(cats)
            ws_overview.add_chart(chart2, 'A' + str(_chart_base + 90))

            # Chart 5 — ROCE bar with dashed 15% target + green average line (VBA step 145)
            chart3 = BarChart()
            chart3.title = 'ROCE'
            chart3.width = 20
            chart3.height = 12
            for r_idx, row in enumerate(consol_rows, start=2):
                if 'ROCE' in row.material_number and 'CAPITAL' not in row.material_number:
                    roce_ref = Reference(ws_overview, min_col=3, max_col=2 + n_periods, min_row=r_idx)
                    chart3.add_data(roce_ref, titles_from_data=False)
                    chart3.series[-1].title = SeriesLabel(v='ROCE')
                    break
            chart3.set_categories(cats)
            target_ref = Reference(ws_overview, min_col=3, max_col=2 + n_periods, min_row=target_row)
            line_overlay = LineChart()
            line_overlay.add_data(target_ref, titles_from_data=False)
            line_overlay.series[0].title = SeriesLabel(v='Target 15%')
            line_overlay.series[0].graphicalProperties.line.dashStyle = 'dash'
            line_overlay.series[0].graphicalProperties.line.solidFill = 'FF0000'
            # Green ROCE average line (VBA lines 3209-3222, RGB(160,208,120)=A0D078, weight 2)
            avg_ref = Reference(ws_overview, min_col=3, max_col=2 + n_periods, min_row=roce_avg_row)
            avg_line = LineChart()
            avg_line.add_data(avg_ref, titles_from_data=False)
            avg_line.series[0].title = SeriesLabel(v='ROCE average')
            avg_line.series[0].graphicalProperties.line.solidFill = 'A0D078'
            avg_line.series[0].graphicalProperties.line.width = 19050  # weight 2 ≈ 1.5pt in EMU
            chart3 += line_overlay
            chart3 += avg_line
            ws_overview.add_chart(chart3, 'A' + str(_chart_base + 114))

            # ---- Top 10 Overstocks sheet (VBA CreateTop10OverstocksChart line 7116) ----
            top10_count = 0
            try:
                # Use passed-in engine; fall back to constructing one from _IQEngine
                _iq_instance = inventory_quality_engine
                if _iq_instance is None and _IQEngine is not None:
                    _iq_instance = _IQEngine(self.data, self.results, self.value_results)
                if _iq_instance is not None:
                    iq_data = _iq_instance.calculate()
                    top10 = iq_data.get('top_10_overstocks', [])
                    t10_periods = iq_data.get('periods', self.data.periods)
                    if top10:
                        ws_t10 = wb.create_sheet('Top 10 overstocks')
                        top10_sorted = sorted(top10, key=lambda x: x.get('starting_overstock', 0), reverse=True)
                        num_mats = len(top10_sorted)
                        num_p = len(t10_periods)

                        # Row 1: headers — col A = 'Period', cols B+ = material names
                        ws_t10.cell(row=1, column=1, value='Period')
                        for mi, item in enumerate(top10_sorted, start=2):
                            ws_t10.cell(row=1, column=mi, value=item.get('material_name') or item['material_number'])

                        # Rows 2+: one row per period
                        for ri, p in enumerate(t10_periods, start=2):
                            ws_t10.cell(row=ri, column=1, value=p)
                            for mi, item in enumerate(top10_sorted, start=2):
                                pdata = item.get('periods', {}).get(p, {})
                                val = pdata.get('overstock', 0) if isinstance(pdata, dict) else 0
                                ws_t10.cell(row=ri, column=mi, value=round(val, 0))
                                ws_t10.cell(row=ri, column=mi).number_format = '#,##0'

                        # Column widths: col A = period labels, cols B+ = material names
                        ws_t10.column_dimensions['A'].width = 12
                        for mi in range(2, num_mats + 2):
                            from openpyxl.utils import get_column_letter
                            ws_t10.column_dimensions[get_column_letter(mi)].width = 18

                        # Chart: vertical stacked (xlColumnStacked), X-axis = periods, series = materials
                        # VBA: CreateTop10OverstocksChart line 7116, legend at bottom (xlLegendPositionBottom)
                        from openpyxl.chart import BarChart, Reference
                        chart_t10 = BarChart()
                        chart_t10.type = 'col'
                        chart_t10.grouping = 'stacked'
                        chart_t10.overlap = 100
                        chart_t10.title = 'Top 10 Overstocks'
                        chart_t10.y_axis.title = 'Value (€)'
                        chart_t10.y_axis.numFmt = '€#,##0'
                        chart_t10.width = 25
                        chart_t10.height = 15
                        chart_t10.legend.position = 'b'

                        max_row = 1 + num_p
                        for col_idx in range(2, num_mats + 2):
                            data_ref = Reference(ws_t10, min_col=col_idx, min_row=1, max_row=max_row)
                            chart_t10.add_data(data_ref, titles_from_data=True)
                        cats_t10 = Reference(ws_t10, min_col=1, min_row=2, max_row=max_row)
                        chart_t10.set_categories(cats_t10)

                        ws_t10.add_chart(chart_t10, 'A' + str(max_row + 2))
                        top10_count = num_mats
            except Exception as e:
                print(f"  Warning: Top 10 overstocks sheet skipped: {e}")

            # ---- Inventory quality chart sheet (VBA CreateInventoryQualityChart line 6925) ----
            iq_chart_done = False
            try:
                # Reuse _iq_instance / iq_data from the Top 10 block above if available
                _iq2 = inventory_quality_engine
                if _iq2 is None and _IQEngine is not None:
                    _iq2 = _IQEngine(self.data, self.results, self.value_results)
                if _iq2 is not None:
                    iq_data2 = _iq2.calculate()
                    period_totals = iq_data2.get('period_totals', {})
                    iq_periods = iq_data2.get('periods', self.data.periods)
                    if period_totals and iq_periods:
                        ws_iq = wb.create_sheet('Inventory quality chart')
                        ws_iq.sheet_properties.tabColor = '833C0C'

                        # Extract COGS from value_results consolidation
                        cogs_row = None
                        for lt, rows in self.value_results.items():
                            for r in rows:
                                if 'COST OF GOODS' in str(r.material_number).upper() or \
                                   'COST OF GOODS' in str(r.material_name).upper():
                                    cogs_row = r
                                    break
                            if cogs_row:
                                break

                        # Write header row
                        headers = ['Period', 'Under', 'Safety Stock', 'Strategic Stock',
                                   'Normal Variation', 'Overstock', 'Actual Stock', 'Cost of Goods']
                        for ci, h in enumerate(headers, start=1):
                            ws_iq.cell(row=1, column=ci, value=h)

                        # Write data rows
                        for ri, p in enumerate(iq_periods, start=2):
                            pt = period_totals.get(p, {})
                            cogs_val = 0.0
                            if cogs_row and p in cogs_row.values:
                                try:
                                    cogs_val = float(cogs_row.values[p] or 0)
                                except (TypeError, ValueError):
                                    cogs_val = 0.0
                            row_vals = [
                                p,
                                round(pt.get('under', 0), 0),
                                round(pt.get('safety', 0), 0),
                                round(pt.get('strategic', 0), 0),
                                round(pt.get('normal', 0), 0),
                                round(pt.get('overstock', 0), 0),
                                round(pt.get('inventory', 0), 0),
                                round(cogs_val, 0),
                            ]
                            for ci, v in enumerate(row_vals, start=1):
                                cell = ws_iq.cell(row=ri, column=ci, value=v)
                                if ci > 1:
                                    cell.number_format = '#,##0'

                        ws_iq.column_dimensions['A'].width = 12
                        for ci in range(2, 9):
                            from openpyxl.utils import get_column_letter
                            ws_iq.column_dimensions[get_column_letter(ci)].width = 16

                        num_iq_p = len(iq_periods)
                        max_iq_row = 1 + num_iq_p

                        # --- Stacked bar chart (5 bands: Under, Safety, Strategic, Normal, Overstock) ---
                        from openpyxl.chart import BarChart, LineChart, Reference, Series
                        from openpyxl.chart.series import SeriesLabel
                        from openpyxl.drawing.fill import PatternFillProperties

                        BAND_COLS = [
                            (2, 'C00000'),  # Under
                            (3, '196B24'),  # Safety Stock
                            (4, 'BE8C00'),  # Strategic Stock
                            (5, 'FFC000'),  # Normal Variation
                            (6, 'FF0000'),  # Overstock
                        ]

                        bar_iq = BarChart()
                        bar_iq.type = 'col'
                        bar_iq.grouping = 'stacked'
                        bar_iq.overlap = 100
                        bar_iq.title = 'Inventory quality'
                        bar_iq.y_axis.title = 'Value (€)'
                        bar_iq.y_axis.numFmt = '€#,##0'
                        bar_iq.width = 30
                        bar_iq.height = 18
                        bar_iq.legend.position = 'b'

                        for col_idx, hex_color in BAND_COLS:
                            data_ref = Reference(ws_iq, min_col=col_idx, min_row=1, max_row=max_iq_row)
                            bar_iq.add_data(data_ref, titles_from_data=True)
                            s = bar_iq.series[-1]
                            s.graphicalProperties.solidFill = hex_color
                            s.graphicalProperties.line.solidFill = hex_color

                        cats_iq = Reference(ws_iq, min_col=1, min_row=2, max_row=max_iq_row)
                        bar_iq.set_categories(cats_iq)

                        # --- Line chart overlay (Actual Stock + CoGS) ---
                        line_iq = LineChart()
                        line_iq.grouping = 'standard'

                        act_ref = Reference(ws_iq, min_col=7, min_row=1, max_row=max_iq_row)
                        line_iq.add_data(act_ref, titles_from_data=True)
                        s_act = line_iq.series[-1]
                        s_act.graphicalProperties.line.solidFill = '800080'
                        s_act.graphicalProperties.line.width = 19050  # 1.5pt in EMU

                        cog_ref = Reference(ws_iq, min_col=8, min_row=1, max_row=max_iq_row)
                        line_iq.add_data(cog_ref, titles_from_data=True)
                        s_cog = line_iq.series[-1]
                        s_cog.graphicalProperties.line.solidFill = 'ADD8E6'
                        s_cog.graphicalProperties.line.width = 38100  # 3pt in EMU

                        # Combine: bar + line
                        bar_iq += line_iq
                        ws_iq.add_chart(bar_iq, 'A' + str(max_iq_row + 2))
                        iq_chart_done = True
            except Exception as e:
                print(f"  Warning: Inventory quality chart sheet skipped: {e}")

            # ---- High-level overview (Phase 2) — charts 2 & 3 (VBA lines 3049-3091) ----
            # Source sheets now exist in the workbook; recreate charts referencing their cells.
            try:
                _ws_iq_ov = wb['Inventory quality chart'] if 'Inventory quality chart' in wb.sheetnames else None
                if _ws_iq_ov is not None:
                    _iq_ov_max = _ws_iq_ov.max_row
                    from openpyxl.chart import BarChart as _BC2, LineChart as _LC2, Reference as _Ref2
                    from openpyxl.chart.series import SeriesLabel as _SL2
                    # Chart 2 — stacked bar + line overlay (same as dedicated sheet)
                    _ov_iq = _BC2()
                    _ov_iq.type = 'col'
                    _ov_iq.grouping = 'stacked'
                    _ov_iq.overlap = 100
                    _ov_iq.title = 'Inventory quality'
                    _ov_iq.y_axis.numFmt = '€#,##0'
                    _ov_iq.width = 30
                    _ov_iq.height = 18
                    _ov_iq.legend.position = 'b'
                    for _ci, _hx in [(2, 'C00000'), (3, '196B24'), (4, 'BE8C00'), (5, 'FFC000'), (6, 'FF0000')]:
                        _r2 = _Ref2(_ws_iq_ov, min_col=_ci, min_row=1, max_row=_iq_ov_max)
                        _ov_iq.add_data(_r2, titles_from_data=True)
                        _s2 = _ov_iq.series[-1]
                        _s2.graphicalProperties.solidFill = _hx
                        _s2.graphicalProperties.line.solidFill = _hx
                    _ov_iq.set_categories(_Ref2(_ws_iq_ov, min_col=1, min_row=2, max_row=_iq_ov_max))
                    _ov_iq_line = _LC2()
                    for _ci, _hx, _lw in [(7, '800080', 19050), (8, 'ADD8E6', 38100)]:
                        _r2 = _Ref2(_ws_iq_ov, min_col=_ci, min_row=1, max_row=_iq_ov_max)
                        _ov_iq_line.add_data(_r2, titles_from_data=True)
                        _sl2 = _ov_iq_line.series[-1]
                        _sl2.graphicalProperties.line.solidFill = _hx
                        _sl2.graphicalProperties.line.width = _lw
                    _ov_iq += _ov_iq_line
                    ws_overview.add_chart(_ov_iq, _ov_chart2_anchor)
            except Exception as _e2:
                print(f"  Warning: Overview inventory quality chart skipped: {_e2}")

            try:
                _ws_t10_ov = wb['Top 10 overstocks'] if 'Top 10 overstocks' in wb.sheetnames else None
                if _ws_t10_ov is not None:
                    _t10_ov_max_r = _ws_t10_ov.max_row
                    _t10_ov_max_c = _ws_t10_ov.max_column
                    from openpyxl.chart import BarChart as _BC3, Reference as _Ref3
                    # Chart 3 — stacked bar, legend=right (VBA xlLegendPositionRight for overview copy)
                    _ov_t10 = _BC3()
                    _ov_t10.type = 'col'
                    _ov_t10.grouping = 'stacked'
                    _ov_t10.overlap = 100
                    _ov_t10.title = 'Top 10 Overstocks'
                    _ov_t10.y_axis.numFmt = '€#,##0'
                    _ov_t10.width = 25
                    _ov_t10.height = 15
                    _ov_t10.legend.position = 'r'
                    for _ci in range(2, _t10_ov_max_c + 1):
                        _r3 = _Ref3(_ws_t10_ov, min_col=_ci, min_row=1, max_row=_t10_ov_max_r)
                        _ov_t10.add_data(_r3, titles_from_data=True)
                    _ov_t10.set_categories(_Ref3(_ws_t10_ov, min_col=1, min_row=2, max_row=_t10_ov_max_r))
                    ws_overview.add_chart(_ov_t10, _ov_chart3_anchor)
            except Exception as _e3:
                print(f"  Warning: Overview Top 10 chart skipped: {_e3}")

            # ===== MoM Comparison sheet + scatter chart =====
            mom_done = False
            if previous_cycle_df is not None:
                try:
                    from modules.mom_comparison_engine import MoMComparisonEngine
                    _mom_eng = MoMComparisonEngine(df_volumes, previous_cycle_df)
                    _mom_df = _mom_eng.calculate()
                    if not _mom_df.empty:
                        _mom_df.to_excel(writer, sheet_name='MoM Comparison', index=False)
                        _ws_mom = writer.sheets['MoM Comparison']

                        # --- scatter chart: Current vs Previous inventory ---
                        _scatter_data = _mom_eng.create_scatter_data()
                        if _scatter_data['materials']:
                            # Write scatter source data to a helper sheet
                            _scat_df = pd.DataFrame({
                                'Material': _scatter_data['materials'],
                                'Previous Inventory': _scatter_data['previous'],
                                'Current Inventory': _scatter_data['current'],
                            })
                            _scat_df.to_excel(writer, sheet_name='MoM Scatter Data', index=False)
                            _ws_scat = writer.sheets['MoM Scatter Data']

                            from openpyxl.chart import ScatterChart as _ScatChart, Reference as _ScatRef, Series as _ScatSeries
                            from openpyxl.chart.marker import Marker as _Marker
                            from openpyxl.drawing.fill import PatternFillProperties as _PFP, ColorChoice as _CC

                            _scat_max_r = _scat_df.shape[0] + 1
                            _sc = _ScatChart()
                            _sc.title = 'MoM Inventory Scatter'
                            _sc.x_axis.title = 'Previous Cycle Inventory'
                            _sc.y_axis.title = 'Current Cycle Inventory'
                            _sc.width = 25
                            _sc.height = 15
                            _sc.style = 13

                            _x_vals = _ScatRef(_ws_scat, min_col=2, min_row=2, max_row=_scat_max_r)
                            _y_vals = _ScatRef(_ws_scat, min_col=3, min_row=2, max_row=_scat_max_r)
                            _series = _ScatSeries(_y_vals, _x_vals, title='Materials')
                            _series.graphicalProperties.line.noFill = True  # no connecting line
                            _sc.series.append(_series)

                            # Apply per-point colours from quadrant logic
                            for _pi, _hex_col in enumerate(_scatter_data['colors']):
                                from openpyxl.chart.series import DataPoint as _DP
                                from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
                                _dp = _DP(idx=_pi)
                                _dp.graphicalProperties.solidFill = _hex_col
                                _series.data_points.append(_dp)

                            _ws_mom.add_chart(_sc, 'J2')
                        mom_done = True
                except Exception as _e_mom:
                    print(f"  Warning: MoM comparison sheet skipped: {_e_mom}")

            # ===== Reorder sheets to match VBA workbook tab order =====
            _desired_order = [
                'Planning sheet',
                'Values_Planning sheet',
                'FTE requirements',
                'Inventory quality chart',
                'Top 10 overstocks',
                'High-level overview',
                'MoM Comparison',
                'MoM Scatter Data',
            ]
            _existing = wb.sheetnames
            _ordered = [s for s in _desired_order if s in _existing]
            _ordered += [s for s in _existing if s not in _ordered]
            wb._sheets = [wb[s] for s in _ordered]

        print(f"\nResults exported to: {output_path}")
        print(f"  - Planning sheet (volumes)")
        print(f"  - Values_Planning sheet (financial)")
        if fte_records:
            print(f"  - FTE requirements ({len(fte_records)} groups)")
        if consol_rows:
            print(f"  - High-level overview ({len(consol_rows)} consolidation rows, 3 charts)")
        if top10_count:
            print(f"  - Top 10 overstocks ({top10_count} materials)")
        if iq_chart_done:
            print(f"  - Inventory quality chart")
        if mom_done:
            print(f"  - MoM Comparison (with scatter chart)")

    def _apply_excel_formatting(self, ws, is_values_sheet=False):
        """Apply VBA-matching cell formatting to a planning worksheet.

        Parameters
        ----------
        is_values_sheet : bool
            When True, financial number format '#,##0.00' is used instead of '#,##0',
            and the Starting stock column also receives this format.

        Rules applied:
        • #,##0 (Planning) / #,##0.00 (Values_Planning) number format on data columns
        • 0.0% format on Line 10 (Utilization rate) data cells
        • Dynamic CF: red (FFC7CE) on L04 < 0, L10 > 100%; orange (FFC896) on L10 < 30%
        • Dynamic CF: purple (C8A2C8) on L09 < 1 (availability is 0.0-1.0 scale)
        • L10 rules only applied to production rows where material starts with 'Z_'
        • Bold font on Line 03 (Total demand) rows
        • Dotted top border between material-group boundaries
        """
        from openpyxl.styles import PatternFill, Font, Border, Side
        from openpyxl.formatting.rule import CellIsRule
        import re

        bold_font   = Font(bold=True)
        # Dynamic CF fills (written as rules, not static fills)
        cf_red    = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        cf_orange = PatternFill(start_color='FFC896', end_color='FFC896', fill_type='solid')
        cf_purple = PatternFill(start_color='C8A2C8', end_color='C8A2C8', fill_type='solid')
        dotted      = Side(border_style='dotted', color='000000')

        def _pf(hex_color):
            return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

        # Per-line-type row fills (VBA extracted_vba.txt lines 5630–5904)
        line_type_colors = {
            '01. Demand forecast':        _pf('E2EFDA'),  # light green
            '02. Dependent demand':       _pf('EDEDED'),  # light grey
            '03. Total demand':           _pf('DAEEF3'),  # light blue
            '04. Inventory':              _pf('FFF2CC'),  # light yellow
            '05. Minimum target stock':   _pf('E2EFDA'),  # light green
            '06. Production plan':        _pf('FCE4D6'),  # light peach
            '06. Purchase receipt':       _pf('FCE4D6'),  # light peach
            '07. Purchase plan':          _pf('D9E1F2'),  # light blue-grey
            '07. Capacity utilization':   _pf('D9E1F2'),  # light blue-grey
            '08. Dependent requirements': _pf('F2F2F2'),  # very light grey
            '09. Available capacity':     _pf('E2EFDA'),  # light green
            '10. Utilization rate':       _pf('FCE4D6'),  # light peach
            '11. Shift availability':     _pf('D9D9D9'),  # medium grey
            '12. FTE requirements':       _pf('FFF2CC'),  # light yellow
        }

        total_cols = ws.max_column
        period_re  = re.compile(r'^\d{4}-\d{2}$')

        num_fmt = '#,##0.00' if is_values_sheet else '#,##0'

        # Locate key columns by reading the header row
        line_type_col = mat_num_col = data_col_start = starting_stock_col = None
        aux_col_idx = aux2_col_idx = None
        for cell in ws[1]:
            hdr = str(cell.value or '')
            if hdr == 'Line type':
                line_type_col = cell.column
            elif hdr == 'Material number':
                mat_num_col = cell.column
            elif hdr == 'Starting stock':
                starting_stock_col = cell.column
            elif hdr == 'Aux Column':
                aux_col_idx = cell.column
            elif hdr == 'Aux 2 Column':
                aux2_col_idx = cell.column
            elif period_re.match(hdr) and data_col_start is None:
                data_col_start = cell.column

        if line_type_col is None:
            return
        if data_col_start is None:
            data_col_start = total_cols + 1  # no period columns found

        # Convert period header text strings to Excel date values with mm/yyyy format
        # (VBA Add_Headers_Planningsheet line 4862: date value + mm/yyyy NumberFormat)
        from datetime import datetime as _dt
        for cell in ws[1]:
            hdr = str(cell.value or '')
            if period_re.match(hdr):
                try:
                    cell.value = _dt.strptime(hdr, '%Y-%m')
                    cell.number_format = 'mm/yyyy'
                except ValueError:
                    pass

        prev_mat = None
        _cf_l04_rows = []
        _cf_l09_rows = []
        _cf_l10_rows = []
        for excel_row in range(2, ws.max_row + 1):
            lt  = ws.cell(row=excel_row, column=line_type_col).value
            mat = ws.cell(row=excel_row, column=mat_num_col).value if mat_num_col else None

            is_l03 = lt == '03. Total demand'
            is_l04 = lt == '04. Inventory'
            is_l09 = lt == '09. Available capacity'
            is_l10 = lt == '10. Utilization rate'

            # Dotted top border at material-group boundaries
            if mat is not None and mat != prev_mat and excel_row > 2:
                for col in range(1, total_cols + 1):
                    c = ws.cell(row=excel_row, column=col)
                    b = c.border
                    c.border = Border(top=dotted, bottom=b.bottom,
                                      left=b.left, right=b.right)
            prev_mat = mat

            # Collect rows for dynamic CF (second pass)
            if is_l04:
                _cf_l04_rows.append(excel_row)
            elif is_l09:
                _cf_l09_rows.append(excel_row)
            elif is_l10 and str(mat or '').startswith('Z_'):
                _cf_l10_rows.append(excel_row)

            # Apply per-line-type row fill to all columns
            row_fill = line_type_colors.get(str(lt or ''))
            if row_fill:
                for _c in range(1, total_cols + 1):
                    ws.cell(row=excel_row, column=_c).fill = row_fill

            for col in range(1, total_cols + 1):
                cell = ws.cell(row=excel_row, column=col)
                is_data = col >= data_col_start

                # Bold for Total demand rows (all columns)
                if is_l03:
                    cell.font = bold_font

                # Apply number format to Starting stock column even though it's not a period
                if starting_stock_col and col == starting_stock_col:
                    cell.number_format = num_fmt

                # VBA FIX 7: Aux columns store exact float, display with #,##0
                # Exception: Line 10 Aux Column uses "0%" (VBA AddVisualsToMaterials)
                if (aux_col_idx and col == aux_col_idx) or \
                        (aux2_col_idx and col == aux2_col_idx):
                    # Coerce string-wrapped numbers to float so number_format works
                    if isinstance(cell.value, str):
                        try:
                            cell.value = float(cell.value)
                        except (ValueError, TypeError):
                            pass
                    if isinstance(cell.value, (int, float)):
                        if is_l10 and aux_col_idx and col == aux_col_idx:
                            cell.number_format = '0%'
                        elif is_l09 and aux2_col_idx and col == aux2_col_idx:
                            cell.number_format = '0%'
                        else:
                            cell.number_format = num_fmt

                if not is_data:
                    continue

                val     = cell.value
                num_val = val if isinstance(val, (int, float)) else None

                # Number format
                if is_l10:
                    cell.number_format = '0.0%'
                elif is_l09:
                    cell.number_format = '0%'
                else:
                    cell.number_format = num_fmt

        # Dynamic conditional formatting (VBA ApplyConditionalFormatting line 4144)
        if data_col_start is not None and data_col_start <= total_cols:
            from openpyxl.utils import get_column_letter as _gcl
            start_letter = _gcl(data_col_start)
            end_letter   = _gcl(total_cols)

            # L10 production rows: red >100%, orange <30% (VBA lines 4189-4198)
            for _rn in _cf_l10_rows:
                _rng = f'{start_letter}{_rn}:{end_letter}{_rn}'
                ws.conditional_formatting.add(_rng, CellIsRule(operator='greaterThan', formula=['1'],   fill=cf_red))
                ws.conditional_formatting.add(_rng, CellIsRule(operator='lessThan',    formula=['0.3'], fill=cf_orange))

            # L04 rows: red <0 (VBA lines 4208-4213)
            for _rn in _cf_l04_rows:
                _rng = f'{start_letter}{_rn}:{end_letter}{_rn}'
                ws.conditional_formatting.add(_rng, CellIsRule(operator='lessThan', formula=['0'], fill=cf_red))

            # L09 rows: purple <1 — availability stored as 0.0-1.0 (VBA lines 4236-4241)
            for _rn in _cf_l09_rows:
                _rng = f'{start_letter}{_rn}:{end_letter}{_rn}'
                ws.conditional_formatting.add(_rng, CellIsRule(operator='lessThan', formula=['1'], fill=cf_purple))

        # ---- Column widths ----
        from openpyxl.utils import get_column_letter as _gcl_w
        _named_widths = {
            'Material number': 18, 'Material name': 22, 'Line type': 22,
            'Product type': 14, 'Product family': 14, 'SPC product': 14,
            'Product cluster': 14, 'Product name': 18,
            'Aux Column': 14, 'Aux 2 Column': 14, 'Starting stock': 14,
        }
        for _hcell in ws[1]:
            _hw = _named_widths.get(str(_hcell.value or ''))
            if _hw:
                ws.column_dimensions[_gcl_w(_hcell.column)].width = _hw
            elif data_col_start is not None and _hcell.column >= data_col_start:
                ws.column_dimensions[_gcl_w(_hcell.column)].width = 12

    def _apply_consol_colors(self, ws):
        """Apply VBA-matching colors and bold to ZZZZZZ_ consolidation rows on Values_Planning sheet."""
        from openpyxl.styles import PatternFill, Font

        def _pf(hex_color):
            return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

        consol_colors = {
            'TURNOVER':                      _pf('A9D08E'),  # RGB(169,208,142) green
            'RAW MATERIAL COST':             _pf('FCE4D6'),  # RGB(252,228,214) light peach
            'MACHINE COST':                  _pf('FCE4D6'),
            'DIRECT FTE COST':               _pf('FCE4D6'),
            'INDIRECT FTE COST':             _pf('FCE4D6'),
            'OVERHEAD COST':                 _pf('FCE4D6'),
            'COST OF GOODS':                 _pf('F8CBAD'),  # RGB(248,203,173) peach
            'GROSS MARGIN':                  _pf('A0D078'),  # RGB(160,208,120) medium green
            'SG&A COST':                     _pf('FCE4D6'),
            'EBITDA':                        _pf('FFE699'),  # RGB(255,230,153) medium yellow
            'D&A COST':                      _pf('FCE4D6'),
            'EBIT':                          _pf('FFE699'),  # RGB(255,230,153) medium yellow
            'FIXED ASSETS NET BOOK VALUE':   _pf('FCE4D6'),  # peach (VBA line 5812+)
            'INVENTORY VALUE':               _pf('FCE4D6'),
            'RECEIVABLES':                   _pf('FCE4D6'),
            'PAYABLES':                      _pf('FCE4D6'),
            'WORKING CAPITAL REQUIREMENTS':  _pf('FFE699'),  # yellow (VBA line 5860+)
            'CAPITAL INVESTMENT':            _pf('FFE699'),
            'OPERATIONAL CASHFLOW':          _pf('FFE699'),
            'ROCE':                          _pf('A9D08E'),  # same green as TURNOVER
        }

        # Rows that receive bold formatting (key summary lines)
        bold_keys = {
            'COST OF GOODS', 'GROSS MARGIN', 'EBITDA', 'EBIT', 'ROCE',
        }

        mat_num_col = None
        for cell in ws[1]:
            if str(cell.value or '') == 'Material number':
                mat_num_col = cell.column
                break
        if mat_num_col is None:
            return

        total_cols = ws.max_column
        for excel_row in range(2, ws.max_row + 1):
            mat_num = ws.cell(row=excel_row, column=mat_num_col).value
            if not mat_num or not str(mat_num).startswith('ZZZZZZ_'):
                continue
            key = str(mat_num).replace('ZZZZZZ_', '')
            fill = consol_colors.get(key)
            is_bold = key in bold_keys
            if fill or is_bold:
                for col in range(1, total_cols + 1):
                    cell = ws.cell(row=excel_row, column=col)
                    if fill:
                        cell.fill = fill
                    if is_bold:
                        # Preserve existing font properties, only set bold
                        existing = cell.font
                        cell.font = Font(
                            bold=True,
                            name=existing.name,
                            size=existing.size,
                            color=existing.color,
                            italic=existing.italic,
                        )

    def _apply_fte_formatting(self, ws, period_cols):
        """Apply formatting to the FTE requirements sheet."""
        from openpyxl.styles import PatternFill, Font, Alignment

        # Row-prefix colours
        blue_fill   = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        orange_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
        purple_fill = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')
        grey_fill   = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
        hdr_fill    = PatternFill(start_color='263238', end_color='263238', fill_type='solid')
        white_font  = Font(color='FFFFFF', bold=True)
        bold_font   = Font(bold=True)

        total_cols = ws.max_column

        # Header row
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal='center')

        # Convert period header strings to Excel date values with mm/yyyy format
        # (matches _apply_excel_formatting() behaviour on Planning and Values_Planning sheets)
        from datetime import datetime as _dt
        for cell in ws[1]:
            hdr = str(cell.value or '')
            if len(hdr) == 7 and hdr.count('-') == 1:
                try:
                    cell.value = _dt.strptime(hdr, '%Y-%m')
                    cell.number_format = 'mm/yyyy'
                except ValueError:
                    pass

        # Column widths
        ws.column_dimensions['A'].width = 20  # Material number
        ws.column_dimensions['B'].width = 28  # Group name
        ws.column_dimensions['C'].width = 14  # FTE needed
        col_letter = 'D'
        for i, _ in enumerate(period_cols):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(4 + i)].width = 12
        # Average column
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(total_cols)].width = 12

        # Data rows
        for excel_row in range(2, ws.max_row + 1):
            mat_cell = ws.cell(row=excel_row, column=1)
            mat_val = str(mat_cell.value or '')

            is_total = mat_val == 'TOTAL'

            # Determine row fill by material_number prefix
            if is_total:
                row_fill = grey_fill
                row_font = bold_font
            elif mat_val.startswith('ZZZZZ'):
                row_fill = purple_fill
                row_font = None
            elif mat_val.startswith('ZZZZ'):
                row_fill = orange_fill
                row_font = None
            elif mat_val.startswith('ZZ'):
                row_fill = blue_fill
                row_font = None
            else:
                row_fill = None
                row_font = None

            for col in range(1, total_cols + 1):
                cell = ws.cell(row=excel_row, column=col)
                if row_fill:
                    cell.fill = row_fill
                if row_font:
                    cell.font = row_font
                # Number format for numeric columns (cols 4+)
                if col >= 4 and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')

    def to_json(self) -> Dict:
        return {
            'summary': self.summary,
            'periods': self.data.periods,
            'results': {
                lt: [row.to_dict() for row in rows]
                for lt, rows in self.results.items()
            }
        }
